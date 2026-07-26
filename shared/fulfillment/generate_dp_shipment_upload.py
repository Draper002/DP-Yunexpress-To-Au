#!/usr/bin/env python3
"""Generate a DropShipZone shipment upload from a carrier order export."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # pragma: no cover - exercised only in an incomplete runtime.
    xlrd = None


YUNEXPRESS_CARRIER = "YunExpress"
SF_INTERNATIONAL_CARRIER = "SF INTERNATIONAL"
DEFAULT_DP_STATUS = "processing"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def clean_identifier(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def date_folder_name(ship_date: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", clean(ship_date))
    if not match:
        return clean(ship_date)
    year, month, day = match.groups()
    return f"{int(year)}年{int(month)}月{int(day)}日"


def default_output_root(ship_date: str) -> Path:
    return Path.home() / "Desktop" / "Codex Folder" / "DropShipZone" / "Codex发货记录" / date_folder_name(ship_date)


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stamp = datetime.now().strftime("%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def read_dp_order_ids(csv_path: Path, status: str = DEFAULT_DP_STATUS) -> set[str]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = [clean(header) for header in (reader.fieldnames or [])]
        reader.fieldnames = fieldnames
        if "Order ID" not in fieldnames:
            raise ValueError("DP order CSV missing Order ID column.")
        rows = [dict(row) for row in reader]

    if "Status" in fieldnames:
        rows = [row for row in rows if clean(row.get("Status")) == status]
    order_ids = [clean(row.get("Order ID")) for row in rows if clean(row.get("Order ID"))]
    if not order_ids:
        suffix = f" with Status={status!r}" if "Status" in fieldnames else ""
        raise ValueError(f"DP order CSV contains no Order ID values{suffix}.")
    duplicate_order_ids = sorted(
        order_id for order_id, count in Counter(order_ids).items() if count > 1
    )
    if duplicate_order_ids:
        raise ValueError(
            "Duplicate Order ID in selected DP rows: "
            + json.dumps(duplicate_order_ids, ensure_ascii=False)
        )
    return set(order_ids)


def read_yunexpress_rows(xlsx_path: Path, prefer_tracking: bool) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        sheet = workbook["订单信息"] if "订单信息" in workbook.sheetnames else workbook.worksheets[0]
        headers = {clean(sheet.cell(1, col).value): col for col in range(1, sheet.max_column + 1)}
        required = ["客户单号", "运单号"]
        missing_headers = [header for header in required if header not in headers]
        if missing_headers:
            raise ValueError(f"YunExpress order file missing columns: {missing_headers}")

        tracking_col = headers.get("跟踪号")
        rows: list[dict[str, str]] = []
        for row_idx in range(2, sheet.max_row + 1):
            order_id = clean(sheet.cell(row_idx, headers["客户单号"]).value)
            waybill = clean(sheet.cell(row_idx, headers["运单号"]).value)
            tracking = clean(sheet.cell(row_idx, tracking_col).value) if tracking_col else ""
            if not order_id and not waybill and not tracking:
                continue
            if not order_id:
                raise ValueError(f"YunExpress row {row_idx} has blank 客户单号.")
            if not waybill:
                raise ValueError(f"YunExpress row {row_idx} has blank 运单号 for {order_id}.")
            tracking_number = tracking if prefer_tracking and tracking else waybill
            if not tracking_number:
                raise ValueError(f"YunExpress row {row_idx} has no usable tracking/waybill for {order_id}.")
            rows.append({
                "Order ID": order_id,
                "Carrier": YUNEXPRESS_CARRIER,
                "Tracking Number": tracking_number,
                "Source Field": "跟踪号" if prefer_tracking and tracking else "运单号",
            })
    finally:
        workbook.close()
    if not rows:
        raise ValueError("No YunExpress shipment rows found.")
    return rows


def read_sf_international_rows(source_path: Path) -> list[dict[str, str]]:
    required = ["客户订单号", "顺丰运单号"]
    rows: list[dict[str, str]] = []

    def append_row(row_idx: int, values: list[object], headers: dict[str, int]) -> None:
        order_id = clean_identifier(values[headers["客户订单号"]])
        waybill = clean_identifier(values[headers["顺丰运单号"]])
        platform_order_col = headers.get("平台订单号")
        platform_order = clean_identifier(values[platform_order_col]) if platform_order_col is not None else ""
        if not order_id and not waybill and not platform_order:
            return
        if not order_id:
            raise ValueError(f"SF International row {row_idx} has blank 客户订单号.")
        if not waybill:
            raise ValueError(f"SF International row {row_idx} has blank 顺丰运单号 for {order_id}.")
        rows.append(
            {
                "Order ID": order_id,
                "Carrier": SF_INTERNATIONAL_CARRIER,
                "Tracking Number": waybill,
                "Source Field": "顺丰运单号",
            }
        )

    suffix = source_path.suffix.lower()
    if suffix == ".xls":
        if xlrd is None:
            raise ValueError("Reading SF International .xls files requires the xlrd package.")
        workbook = xlrd.open_workbook(str(source_path), on_demand=True)
        try:
            sheet = workbook.sheet_by_index(0)
            headers = {clean(sheet.cell_value(0, col)): col for col in range(sheet.ncols)}
            missing_headers = [header for header in required if header not in headers]
            if missing_headers:
                raise ValueError(f"SF International order file missing columns: {missing_headers}")
            for row_idx in range(1, sheet.nrows):
                append_row(row_idx + 1, [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)], headers)
        finally:
            workbook.release_resources()
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            headers = {clean(cell.value): col for col, cell in enumerate(sheet[1])}
            missing_headers = [header for header in required if header not in headers]
            if missing_headers:
                raise ValueError(f"SF International order file missing columns: {missing_headers}")
            for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                append_row(row_idx, list(values), headers)
        finally:
            workbook.close()
    else:
        raise ValueError("SF International order file must be .xls, .xlsx, or .xlsm.")

    if not rows:
        raise ValueError("No SF International shipment rows found.")
    return rows


def validate_rows(
    rows: list[dict[str, str]],
    expected_order_ids: set[str] | None,
    source_label: str,
    expected_carrier: str,
) -> None:
    order_ids = [row["Order ID"] for row in rows]
    tracking_numbers = [row["Tracking Number"] for row in rows]
    duplicate_order_ids = sorted(order_id for order_id, count in Counter(order_ids).items() if count > 1)
    duplicate_tracking = sorted(value for value, count in Counter(tracking_numbers).items() if count > 1)
    if duplicate_order_ids:
        raise ValueError(f"Duplicate Order ID in {source_label} rows: " + json.dumps(duplicate_order_ids, ensure_ascii=False))
    if duplicate_tracking:
        raise ValueError(
            f"Duplicate tracking/waybill in {source_label} rows: "
            + json.dumps(duplicate_tracking, ensure_ascii=False)
        )
    carriers = {row["Carrier"] for row in rows}
    if carriers != {expected_carrier}:
        raise ValueError(f"Unexpected carrier values in {source_label} rows: {sorted(carriers)}")
    if expected_order_ids is not None and set(order_ids) != expected_order_ids:
        missing = sorted(expected_order_ids - set(order_ids))
        extra = sorted(set(order_ids) - expected_order_ids)
        raise ValueError(
            f"{source_label} Order ID set does not match DP CSV. "
            + json.dumps({"missing": missing, "extra": extra}, ensure_ascii=False)
        )


def write_upload(
    template_path: Path,
    rows: list[dict[str, str]],
    date: str,
    output_root: Path,
    carrier: str,
) -> Path:
    workbook = openpyxl.load_workbook(template_path)
    try:
        sheet = workbook.worksheets[0]
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        headers = ["Order ID", "Carrier", "Tracking Number"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(1, col_idx).value = header
        for column, width in {"A": 18, "B": 22, "C": 24}.items():
            sheet.column_dimensions[column].width = max(sheet.column_dimensions[column].width or 0, width)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row_idx, col_idx)
                cell.value = row[header]
                cell.number_format = "@"

        output_dir = output_root / "DP回填模板"
        output_dir.mkdir(parents=True, exist_ok=True)
        carrier_label = "YunExpress" if carrier == YUNEXPRESS_CARRIER else "SF-INTERNATIONAL"
        output_path = unique_path(output_dir / f"DP_发货回填_{date}_{len(rows)}orders_{carrier_label}_waybill.xlsx")
        workbook.save(output_path)
    finally:
        workbook.close()
    return output_path


def validate_output(path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        headers = [sheet.cell(1, col).value for col in range(1, 4)]
        rows = [
            [sheet.cell(row, col).value for col in range(1, 4)]
            for row in range(2, sheet.max_row + 1)
            if any(sheet.cell(row, col).value not in (None, "") for col in range(1, 4))
        ]
        blank_cells = [
            (idx + 2, col + 1)
            for idx, row in enumerate(rows)
            for col, value in enumerate(row)
            if value in (None, "")
        ]
        return {
            "headers_ok": headers == ["Order ID", "Carrier", "Tracking Number"],
            "row_count": len(rows),
            "blank_cells_count": len(blank_cells),
            "carrier_counts": dict(Counter(row[1] for row in rows)),
            "tracking_unique_count": len({row[2] for row in rows}),
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate DP shipment upload from a carrier order export.")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--yunexpress-xlsx", type=Path)
    source.add_argument("--sf-international-file", type=Path)
    parser.add_argument("--shipment-template-xlsx", required=True, type=Path)
    parser.add_argument("--dp-orders-csv", type=Path, default=None)
    parser.add_argument("--date", required=True)
    parser.add_argument("--prefer-tracking", action="store_true", help="Use 跟踪号 when present; default uses 运单号.")
    parser.add_argument("--output-root", type=Path, default=None)
    args = parser.parse_args()

    try:
        output_root = args.output_root or default_output_root(args.date)
        expected_order_ids = read_dp_order_ids(args.dp_orders_csv) if args.dp_orders_csv else None
        if args.sf_international_file:
            source_label = "SF International"
            carrier = SF_INTERNATIONAL_CARRIER
            rows = read_sf_international_rows(args.sf_international_file)
        else:
            source_label = "YunExpress"
            carrier = YUNEXPRESS_CARRIER
            rows = read_yunexpress_rows(args.yunexpress_xlsx, args.prefer_tracking)
        validate_rows(rows, expected_order_ids, source_label, carrier)
        output_path = write_upload(args.shipment_template_xlsx, rows, args.date, output_root, carrier)
        validation = validate_output(output_path)
        if (
            not validation["headers_ok"]
            or validation["blank_cells_count"] != 0
            or validation["carrier_counts"] != {carrier: len(rows)}
            or validation["tracking_unique_count"] != len(rows)
        ):
            raise ValueError(f"Generated DP upload failed validation: {validation}")
        report = {
            "status": "OK",
            "source_platform": source_label,
            "source_tracking_field": sorted({row["Source Field"] for row in rows}),
            "output_path": str(output_path),
            "validation": validation,
        }
        output_path.with_suffix(".report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"OK: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
