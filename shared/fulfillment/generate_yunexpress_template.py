#!/usr/bin/env python3
"""Generate a YunExpress batch shipment template from DropShipZone AU orders."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl


DEFAULT_PRODUCT_CODE = "THPHR"
DEFAULT_COUNTRY = "AU"
DEFAULT_CURRENCY = "USD"
DEFAULT_STATUS = "processing"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_number(value: object) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", clean(value))
    return float(match.group(0)) if match else None


def round3(value: float) -> float:
    return round(value + 1e-12, 3)


def sku_info_score(info: dict[str, object]) -> int:
    fields = ("chinese_name", "english_name", "price_usd", "unit_weight_kg")
    return sum(1 for field in fields if info.get(field) not in (None, ""))


def header_columns(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        header = clean(sheet.cell(1, col).value)
        if header and header not in headers:
            headers[header] = col
    return headers


def date_folder_name(ship_date: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", clean(ship_date))
    if not match:
        return clean(ship_date)
    year, month, day = match.groups()
    return f"{int(year)}年{int(month)}月{int(day)}日"


def default_output_root(ship_date: str) -> Path:
    return Path.home() / "Desktop" / "Codex Folder" / "DropShipZone" / "Codex发货记录" / date_folder_name(ship_date)


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stamp = datetime.now().strftime("%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def read_dp_orders(csv_path: Path, status: str) -> tuple[list[dict[str, str]], dict[str, int]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = [clean(header) for header in (reader.fieldnames or [])]
        reader.fieldnames = fieldnames
        required_headers = {
            "Order ID",
            "Status",
            "Ship to Name",
            "Street",
            "Suburb",
            "State",
            "Postcode",
            "Phone Number",
            "Item Sku",
            "Item Quantity",
        }
        missing_headers = required_headers - set(fieldnames)
        if missing_headers:
            raise ValueError(
                "DP order CSV missing columns: "
                + json.dumps(sorted(missing_headers), ensure_ascii=False)
                + f"; file={csv_path}; detected_columns="
                + json.dumps(fieldnames, ensure_ascii=False)
            )
        all_rows = [dict(row) for row in reader]

    selected = [row for row in all_rows if clean(row.get("Status")) == status]
    if not selected:
        raise ValueError(f"No orders with Status={status!r}.")

    order_ids = [clean(row.get("Order ID")) for row in selected]
    duplicate_order_ids = sorted(order_id for order_id, count in Counter(order_ids).items() if count > 1)
    if duplicate_order_ids:
        raise ValueError(
            "Duplicate Order ID found in selected rows; multi-SKU orders are not automatic yet: "
            + json.dumps(duplicate_order_ids, ensure_ascii=False)
        )

    required_fields = [
        "Order ID",
        "Ship to Name",
        "Street",
        "Suburb",
        "State",
        "Postcode",
        "Phone Number",
        "Item Sku",
        "Item Quantity",
    ]
    missing_values: list[dict[str, str]] = []
    for row in selected:
        order_id = clean(row.get("Order ID"))
        for field in required_fields:
            if not clean(row.get(field)):
                missing_values.append({"order_id": order_id, "field": field})
    if missing_values:
        raise ValueError("Selected DP orders have blank required fields: " + json.dumps(missing_values, ensure_ascii=False))

    return selected, {
        "total_rows": len(all_rows),
        "selected_rows": len(selected),
        "skipped_rows": len(all_rows) - len(selected),
    }


def read_sku_info(xlsx_path: Path) -> dict[str, dict[str, object]]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook.worksheets[0]
    headers = header_columns(sheet)

    zh_col = headers.get("申报中文名", 2)
    en_col = headers.get("申报英文名", 3)
    sku_col = headers.get("SKU", 4)
    price_col = headers.get("单价", 5)
    weight_col = headers.get("单重", 6)

    sku_map: dict[str, dict[str, object]] = {}
    for row in range(2, sheet.max_row + 1):
        sku = clean(sheet.cell(row, sku_col).value)
        if not sku:
            continue
        candidate = {
            "chinese_name": clean(sheet.cell(row, zh_col).value),
            "english_name": clean(sheet.cell(row, en_col).value),
            "price_usd": parse_number(sheet.cell(row, price_col).value),
            "unit_weight_kg": parse_number(sheet.cell(row, weight_col).value),
        }
        current = sku_map.get(sku)
        if current is None or sku_info_score(candidate) > sku_info_score(current):
            sku_map[sku] = candidate
    return sku_map


def set_value(row_values: list[object], header_to_index: dict[str, int], header: str, value: object) -> None:
    if header not in header_to_index:
        raise ValueError(f"YunExpress template missing header: {header}")
    row_values[header_to_index[header]] = value


def build_yunexpress_rows(
    orders: list[dict[str, str]],
    sku_map: dict[str, dict[str, object]],
    template_headers: list[str],
    product_code: str,
) -> list[list[object]]:
    header_to_index = {header: idx for idx, header in enumerate(template_headers)}
    output_rows: list[list[object]] = []
    missing_skus: list[str] = []

    for order in orders:
        sku = clean(order.get("Item Sku"))
        info = sku_map.get(sku)
        if (
            not info
            or not info.get("chinese_name")
            or not info.get("english_name")
            or info.get("price_usd") is None
            or info.get("unit_weight_kg") is None
        ):
            missing_skus.append(sku)
            continue

        quantity = parse_number(order.get("Item Quantity"))
        if not quantity or quantity <= 0:
            raise ValueError(f"Invalid quantity for Order ID {clean(order.get('Order ID'))}: {order.get('Item Quantity')}")

        row_values: list[object] = [None] * len(template_headers)
        unit_weight = float(info["unit_weight_kg"])
        set_value(row_values, header_to_index, "客户订单号", clean(order.get("Order ID")))
        set_value(row_values, header_to_index, "产品代码", product_code)
        set_value(row_values, header_to_index, "收件人国家", DEFAULT_COUNTRY)
        set_value(row_values, header_to_index, "收件人姓名", clean(order.get("Ship to Name")))
        set_value(row_values, header_to_index, "收件人地址", clean(order.get("Street")))
        set_value(row_values, header_to_index, "收件人城市", clean(order.get("Suburb")))
        set_value(row_values, header_to_index, "收件人省州", clean(order.get("State")))
        set_value(row_values, header_to_index, "邮编", clean(order.get("Postcode")))
        set_value(row_values, header_to_index, "电话", clean(order.get("Phone Number")))
        set_value(row_values, header_to_index, "件数", 1)
        set_value(row_values, header_to_index, "包裹总重量", round3(unit_weight * quantity))
        set_value(row_values, header_to_index, "申报币种", DEFAULT_CURRENCY)
        set_value(row_values, header_to_index, "SKU1", sku)
        set_value(row_values, header_to_index, "申报品名1", info["english_name"])
        set_value(row_values, header_to_index, "中文申报品名1", info["chinese_name"])
        set_value(row_values, header_to_index, "申报数量1", quantity)
        set_value(row_values, header_to_index, "申报FOB价1", info["price_usd"])
        set_value(row_values, header_to_index, "申报单重1", unit_weight)
        output_rows.append(row_values)

    if missing_skus:
        raise ValueError("Missing or incomplete SKU declaration data: " + json.dumps(sorted(set(missing_skus)), ensure_ascii=False))
    return output_rows


def write_template(template_path: Path, rows: list[list[object]], date: str, product_code: str, output_root: Path) -> Path:
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.worksheets[0]
    headers = [clean(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row_idx, col_idx).value = value

    output_dir = output_root / "云途上传模板"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = unique_path(output_dir / f"云途批量寄件_{date}_{len(rows)}orders_{product_code}.xlsx")
    workbook.save(output_path)
    return output_path


def validate_output(path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.worksheets[0]
    rows = [row for row in range(2, sheet.max_row + 1) if clean(sheet.cell(row, 1).value)]
    required_cols = [1, 2, 14, 15, 18, 19, 20, 21, 22, 27, 47, 48, 50, 51, 53]
    missing = [(row, col) for row in rows for col in required_cols if sheet.cell(row, col).value in (None, "")]
    return {
        "row_count": len(rows),
        "required_missing_count": len(missing),
        "product_codes": dict(Counter(clean(sheet.cell(row, 2).value) for row in rows)),
        "countries": dict(Counter(clean(sheet.cell(row, 14).value) for row in rows)),
        "sku_counts": dict(Counter(clean(sheet.cell(row, 47).value) for row in rows)),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate YunExpress batch template from DP orders.")
    parser.add_argument("--dp-orders-csv", required=True, type=Path)
    parser.add_argument("--sku-xlsx", required=True, type=Path)
    parser.add_argument("--yunexpress-template-xlsx", required=True, type=Path)
    parser.add_argument("--date", required=True)
    parser.add_argument("--product-code", default=DEFAULT_PRODUCT_CODE)
    parser.add_argument("--status", default=DEFAULT_STATUS)
    parser.add_argument("--output-root", type=Path, default=None)
    args = parser.parse_args()

    try:
        output_root = args.output_root or default_output_root(args.date)
        orders, source_counts = read_dp_orders(args.dp_orders_csv, args.status)
        sku_map = read_sku_info(args.sku_xlsx)
        template_wb = openpyxl.load_workbook(args.yunexpress_template_xlsx, read_only=True)
        template_sheet = template_wb.worksheets[0]
        template_headers = [clean(template_sheet.cell(1, col).value) for col in range(1, template_sheet.max_column + 1)]
        template_wb.close()
        output_rows = build_yunexpress_rows(orders, sku_map, template_headers, args.product_code)
        output_path = write_template(args.yunexpress_template_xlsx, output_rows, args.date, args.product_code, output_root)
        validation = validate_output(output_path)
        if validation["required_missing_count"] != 0:
            raise ValueError(f"Generated file has missing required cells: {validation}")
        report = {
            "status": "OK",
            "source_counts": source_counts,
            "output_path": str(output_path),
            "validation": validation,
        }
        report_path = output_path.with_suffix(".report.json")
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"OK: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

