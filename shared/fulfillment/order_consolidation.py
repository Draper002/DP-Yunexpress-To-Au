#!/usr/bin/env python3
"""Build and validate same-recipient, same-SKU parcel consolidation plans."""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


PLAN_VERSION = 1
FEATURE_NAME = "same_recipient_same_sku"
RECIPIENT_FIELDS = (
    "Ship to Name",
    "Street",
    "Suburb",
    "State",
    "Postcode",
    "Phone Number",
)


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_quantity(value: object) -> float:
    text = clean(value).replace(",", "")
    try:
        quantity = float(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid item quantity: {value!r}") from exc
    if quantity <= 0:
        raise ValueError(f"Item quantity must be greater than zero: {value!r}")
    return quantity


def display_quantity(value: float) -> int | float:
    return int(value) if float(value).is_integer() else round(float(value), 6)


def filename_quantity(value: float) -> str:
    displayed = display_quantity(value)
    return str(displayed).replace(".", "-")


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).casefold()
    return re.sub(r"\s+", " ", text).strip()


def normalize_phone(value: object) -> str:
    return "".join(character for character in unicodedata.normalize("NFKC", clean(value)) if character.isdigit())


def recipient_key(order: Mapping[str, object]) -> tuple[str, ...]:
    values = [normalize_text(order.get(field)) for field in RECIPIENT_FIELDS[:-1]]
    values.append(normalize_phone(order.get("Phone Number")))
    return tuple(values)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def candidate_groups(orders: list[dict[str, str]]) -> list[list[dict[str, str]]]:
    grouped: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    for order in orders:
        sku = clean(order.get("Item Sku"))
        grouped[recipient_key(order) + (sku,)].append(order)
    return [rows for rows in grouped.values() if len(rows) > 1]


def candidate_summary(orders: list[dict[str, str]]) -> dict[str, int]:
    groups = candidate_groups(orders)
    return {
        "merged_group_count": len(groups),
        "merged_order_count": sum(len(group) for group in groups),
        "saved_parcel_count": sum(len(group) - 1 for group in groups),
    }


def _merge_reference(ship_date: str, index: int, existing_order_ids: set[str]) -> str:
    date_match = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", ship_date)
    date_part = "".join(date_match.groups())[2:] if date_match else re.sub(r"\D", "", ship_date)[-6:]
    date_part = date_part or datetime.now().strftime("%y%m%d")
    candidate_index = index
    while True:
        candidate = f"M{date_part}{candidate_index:03d}"
        if candidate.casefold() not in existing_order_ids:
            return candidate
        candidate_index += 1


def build_plan(
    orders: list[dict[str, str]],
    ship_date: str,
    platform: str,
    source_csv: Path,
    sku_map: Mapping[str, Mapping[str, object]] | None = None,
    merge_enabled: bool = True,
) -> dict[str, object]:
    if not orders:
        raise ValueError("Cannot build a consolidation plan without DP orders.")

    order_ids = [clean(order.get("Order ID")) for order in orders]
    duplicate_order_ids = sorted(order_id for order_id, count in Counter(order_ids).items() if count > 1)
    if not all(order_ids):
        raise ValueError("DP orders contain a blank Order ID.")
    if duplicate_order_ids:
        raise ValueError("Duplicate Order ID in consolidation source: " + json.dumps(duplicate_order_ids, ensure_ascii=False))

    for order in orders:
        missing = [field for field in (*RECIPIENT_FIELDS, "Item Sku", "Item Quantity") if not clean(order.get(field))]
        if missing:
            raise ValueError(
                f"DP Order {clean(order.get('Order ID'))} is missing consolidation fields: "
                + json.dumps(missing, ensure_ascii=False)
            )

    grouped_indices: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for index, order in enumerate(orders):
        grouped_indices[recipient_key(order) + (clean(order.get("Item Sku")),)].append(index)

    existing_ids = {value.casefold() for value in order_ids}
    consumed: set[int] = set()
    packages: list[dict[str, object]] = []
    merge_index = 1

    for index, order in enumerate(orders):
        if index in consumed:
            continue
        group_indices = grouped_indices[recipient_key(order) + (clean(order.get("Item Sku")),)]
        should_merge = merge_enabled and len(group_indices) > 1
        selected_indices = group_indices if should_merge else [index]
        consumed.update(selected_indices)
        selected_orders = [orders[item] for item in selected_indices]
        sku = clean(order.get("Item Sku"))
        quantities = [parse_quantity(item.get("Item Quantity")) for item in selected_orders]
        reference = clean(order.get("Order ID"))
        if should_merge:
            reference = _merge_reference(ship_date, merge_index, existing_ids)
            existing_ids.add(reference.casefold())
            merge_index += 1
        sku_info = (sku_map or {}).get(sku, {})
        packages.append(
            {
                "carrier_reference": reference,
                "merged": should_merge,
                "order_count": len(selected_orders),
                "sku": sku,
                "chinese_name": clean(sku_info.get("chinese_name")),
                "english_name": clean(sku_info.get("english_name")),
                "total_quantity": display_quantity(sum(quantities)),
                "recipient": {field: clean(order.get(field)) for field in RECIPIENT_FIELDS},
                "order_lines": [
                    {
                        "order_id": clean(item.get("Order ID")),
                        "quantity": display_quantity(quantity),
                    }
                    for item, quantity in zip(selected_orders, quantities)
                ],
            }
        )

    merged_packages = [package for package in packages if package["merged"]]
    plan = {
        "version": PLAN_VERSION,
        "feature": FEATURE_NAME,
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "ship_date": ship_date,
        "platform": platform,
        "source_csv": source_csv.name,
        "source_sha256": sha256_file(source_csv),
        "merge_enabled": bool(merge_enabled),
        "order_count": len(orders),
        "package_count": len(packages),
        "merged_group_count": len(merged_packages),
        "merged_order_count": sum(int(package["order_count"]) for package in merged_packages),
        "saved_parcel_count": len(orders) - len(packages),
        "packages": packages,
    }
    validate_plan(plan)
    return plan


def validate_plan(plan: Mapping[str, object]) -> None:
    if plan.get("version") != PLAN_VERSION or plan.get("feature") != FEATURE_NAME:
        raise ValueError("Unsupported consolidation plan version or feature.")
    packages = plan.get("packages")
    if not isinstance(packages, list) or not packages:
        raise ValueError("Consolidation plan contains no packages.")

    references: list[str] = []
    order_ids: list[str] = []
    merged_group_count = 0
    for package in packages:
        if not isinstance(package, dict):
            raise ValueError("Consolidation plan package is not an object.")
        reference = clean(package.get("carrier_reference"))
        sku = clean(package.get("sku"))
        lines = package.get("order_lines")
        if not reference or not sku or not isinstance(lines, list) or not lines:
            raise ValueError("Consolidation plan package is incomplete.")
        references.append(reference)
        line_order_ids = [clean(line.get("order_id")) for line in lines if isinstance(line, dict)]
        if len(line_order_ids) != len(lines) or not all(line_order_ids):
            raise ValueError(f"Consolidation package {reference} contains an invalid order line.")
        order_ids.extend(line_order_ids)
        quantity_total = sum(parse_quantity(line.get("quantity")) for line in lines)
        if abs(quantity_total - parse_quantity(package.get("total_quantity"))) > 1e-9:
            raise ValueError(f"Consolidation package {reference} has an inconsistent quantity total.")
        is_merged = bool(package.get("merged"))
        if is_merged:
            merged_group_count += 1
            if len(lines) < 2:
                raise ValueError(f"Merged package {reference} must contain at least two orders.")
        elif len(lines) != 1 or reference != line_order_ids[0]:
            raise ValueError(f"Non-merged package {reference} must map to exactly its original order.")

    duplicate_refs = sorted(value for value, count in Counter(references).items() if count > 1)
    duplicate_orders = sorted(value for value, count in Counter(order_ids).items() if count > 1)
    if duplicate_refs:
        raise ValueError("Duplicate carrier references in consolidation plan: " + json.dumps(duplicate_refs, ensure_ascii=False))
    if duplicate_orders:
        raise ValueError("DP orders assigned to multiple consolidation packages: " + json.dumps(duplicate_orders, ensure_ascii=False))
    if int(plan.get("order_count", -1)) != len(order_ids):
        raise ValueError("Consolidation plan order count is inconsistent.")
    if int(plan.get("package_count", -1)) != len(packages):
        raise ValueError("Consolidation plan package count is inconsistent.")
    if int(plan.get("merged_group_count", -1)) != merged_group_count:
        raise ValueError("Consolidation plan merged group count is inconsistent.")


def load_plan(
    path: Path,
    source_csv: Path | None = None,
    expected_platform: str | None = None,
    expected_date: str | None = None,
) -> dict[str, object]:
    plan = json.loads(path.read_text(encoding="utf-8"))
    validate_plan(plan)
    if source_csv and plan.get("source_sha256") != sha256_file(source_csv):
        raise ValueError("Consolidation plan does not belong to the selected DP order CSV.")
    if expected_platform and plan.get("platform") != expected_platform:
        raise ValueError(
            f"Consolidation plan platform mismatch: expected {expected_platform}, found {plan.get('platform')}"
        )
    if expected_date and plan.get("ship_date") != expected_date:
        raise ValueError(
            f"Consolidation plan date mismatch: expected {expected_date}, found {plan.get('ship_date')}"
        )
    return plan


def package_by_reference(plan: Mapping[str, object]) -> dict[str, dict[str, object]]:
    return {
        clean(package["carrier_reference"]): package
        for package in plan["packages"]  # type: ignore[index]
    }


def order_to_package(plan: Mapping[str, object]) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for package in plan["packages"]:  # type: ignore[index]
        for line in package["order_lines"]:
            result[clean(line["order_id"])] = package
    return result


def consolidated_orders(orders: list[dict[str, str]], plan: Mapping[str, object]) -> list[dict[str, str]]:
    source_by_id = {clean(order.get("Order ID")): order for order in orders}
    planned_order_ids = set(order_to_package(plan))
    if set(source_by_id) != planned_order_ids:
        raise ValueError(
            "Consolidation plan order set does not match DP orders: "
            + json.dumps(
                {
                    "missing": sorted(set(source_by_id) - planned_order_ids),
                    "extra": sorted(planned_order_ids - set(source_by_id)),
                },
                ensure_ascii=False,
            )
        )
    result: list[dict[str, str]] = []
    for package in plan["packages"]:  # type: ignore[index]
        first_order_id = clean(package["order_lines"][0]["order_id"])
        row = dict(source_by_id[first_order_id])
        row["Order ID"] = clean(package["carrier_reference"])
        row["Item Quantity"] = str(display_quantity(float(package["total_quantity"])))
        result.append(row)
    return result


def validate_carrier_rows(carrier_rows: list[dict[str, str]], plan: Mapping[str, object]) -> None:
    references = [clean(row.get("Order ID")) for row in carrier_rows]
    tracking_numbers = [clean(row.get("Tracking Number")) for row in carrier_rows]
    duplicate_references = sorted(value for value, count in Counter(references).items() if count > 1)
    duplicate_tracking = sorted(value for value, count in Counter(tracking_numbers).items() if count > 1)
    if duplicate_references:
        raise ValueError("Duplicate carrier references in carrier export: " + json.dumps(duplicate_references, ensure_ascii=False))
    if duplicate_tracking:
        raise ValueError("Duplicate tracking numbers between carrier packages: " + json.dumps(duplicate_tracking, ensure_ascii=False))
    expected = set(package_by_reference(plan))
    actual = set(references)
    if actual != expected:
        raise ValueError(
            "Carrier reference set does not match consolidation plan. "
            + json.dumps({"missing": sorted(expected - actual), "extra": sorted(actual - expected)}, ensure_ascii=False)
        )


def expand_carrier_rows(
    carrier_rows: list[dict[str, str]],
    plan: Mapping[str, object],
) -> list[dict[str, str]]:
    validate_carrier_rows(carrier_rows, plan)
    packages = package_by_reference(plan)
    expanded: list[dict[str, str]] = []
    for carrier_row in carrier_rows:
        reference = clean(carrier_row.get("Order ID"))
        package = packages[reference]
        for line in package["order_lines"]:
            row = dict(carrier_row)
            row["Order ID"] = clean(line["order_id"])
            row["Carrier Reference"] = reference
            row["Merged Group"] = reference if package["merged"] else ""
            expanded.append(row)
    return expanded


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="DCEBFA")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="16324F")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 12,
        "B": 20,
        "C": 22,
        "D": 18,
        "E": 20,
        "F": 14,
        "G": 14,
        "H": 12,
        "I": 20,
        "J": 18,
        "K": 42,
        "L": 16,
        "M": 12,
        "N": 10,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_plan_files(plan: Mapping[str, object], carrier_output: Path) -> tuple[Path, Path]:
    stem = f"订单合并关系_{carrier_output.stem}"
    json_path = carrier_output.parent / f"{stem}.json"
    workbook_path = carrier_output.parent / f"{stem}.xlsx"
    json_path.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "批次概览"
    summary.append(["项目", "值"])
    summary_rows = [
        ("发货日期", plan["ship_date"]),
        ("承运平台", plan["platform"]),
        ("原始DP订单数", plan["order_count"]),
        ("快递包裹数", plan["package_count"]),
        ("合并组数", plan["merged_group_count"]),
        ("涉及合并订单数", plan["merged_order_count"]),
        ("节省包裹数", plan["saved_parcel_count"]),
        ("合并规则", "收件人六项标准化后一致 + SKU完全一致"),
    ]
    for row in summary_rows:
        summary.append(row)
    _style_sheet(summary)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 56

    details = workbook.create_sheet("订单关系")
    details.append(
        [
            "是否合并",
            "承运商客户单号",
            "DP订单号",
            "SKU",
            "申报中文名",
            "原订单数量",
            "合并总数量",
            "组内订单数",
            "收件人姓名",
            "电话",
            "地址",
            "城市",
            "州",
            "邮编",
        ]
    )
    for package in plan["packages"]:  # type: ignore[index]
        recipient = package["recipient"]
        for line in package["order_lines"]:
            details.append(
                [
                    "是" if package["merged"] else "否",
                    package["carrier_reference"],
                    line["order_id"],
                    package["sku"],
                    package["chinese_name"],
                    line["quantity"],
                    package["total_quantity"],
                    package["order_count"],
                    recipient["Ship to Name"],
                    recipient["Phone Number"],
                    recipient["Street"],
                    recipient["Suburb"],
                    recipient["State"],
                    recipient["Postcode"],
                ]
            )
    _style_sheet(details)
    workbook.save(workbook_path)
    workbook.close()
    return json_path, workbook_path


def carrier_tracking_by_reference(carrier_rows: Iterable[Mapping[str, object]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in carrier_rows:
        reference = clean(row.get("Order ID"))
        tracking = clean(row.get("Tracking Number"))
        if reference:
            result[reference] = tracking
    return result


def write_tracking_workbook(
    plan: Mapping[str, object],
    carrier_rows: list[dict[str, str]],
    output_file: Path,
    exact_output_path: bool = False,
) -> Path | None:
    merged_packages = [package for package in plan["packages"] if package["merged"]]  # type: ignore[index]
    if not merged_packages:
        return None
    tracking_by_reference = carrier_tracking_by_reference(carrier_rows)
    output_path = (
        output_file
        if exact_output_path
        else output_file.parent / f"合并订单与运单对应表_{output_file.stem}.xlsx"
    )
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "合并订单运单关系"
    sheet.append(
        [
            "运单号",
            "承运商客户单号",
            "DP订单号",
            "SKU",
            "申报中文名",
            "原订单数量",
            "合并总数量",
            "组内订单数",
        ]
    )
    for package in merged_packages:
        reference = clean(package["carrier_reference"])
        for line in package["order_lines"]:
            sheet.append(
                [
                    tracking_by_reference[reference],
                    reference,
                    line["order_id"],
                    package["sku"],
                    package["chinese_name"],
                    line["quantity"],
                    package["total_quantity"],
                    package["order_count"],
                ]
            )
    _style_sheet(sheet)
    workbook.save(output_path)
    workbook.close()
    return output_path


def write_supplier_instruction(
    package: Mapping[str, object],
    tracking_number: str,
    output_path: Path,
    sender_name: str = "",
) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "合并件说明"
    sheet.append(
        [
            "运单号",
            "合并编号",
            "DP订单号",
            "SKU",
            "申报中文名",
            "原订单数量",
            "本包裹合计数量",
            "组内订单数",
            "寄方姓名",
            "供应商操作要求",
        ]
    )
    for line in package["order_lines"]:  # type: ignore[index]
        sheet.append(
            [
                tracking_number,
                package["carrier_reference"],
                line["order_id"],
                package["sku"],
                package["chinese_name"],
                line["quantity"],
                package["total_quantity"],
                package["order_count"],
                sender_name,
                f"请将以上 {package['order_count']} 个订单合并装入一个包裹，SKU {package['sku']} 合计 {package['total_quantity']} 件，仅使用本文件夹中的一张面单。",
            ]
        )
    _style_sheet(sheet)
    sheet.column_dimensions["J"].width = 72
    for row in sheet.iter_rows(min_row=2, min_col=10, max_col=10):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(output_path)
    workbook.close()
    return output_path
