#!/usr/bin/env python3
"""
Sort YunExpress label PDFs from one ZIP into SKU folders.

The matching chain is:
YunExpress PDF filename -> YunExpress waybill no. -> DP Order ID -> SKU -> declaration Chinese name.

The script is strict by default. If any PDF cannot be matched, or if an expected
YunExpress waybill has no PDF, it stops before producing a "success" package.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl


WAYBILL_RE = re.compile(r"(YT[A-Za-z0-9]+)", re.IGNORECASE)


@dataclass(frozen=True)
class Shipment:
    order_id: str
    waybill: str
    tracking: str
    status: str


@dataclass(frozen=True)
class LabelPdf:
    zip_name: str
    waybill: str


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def header_columns(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        header = clean(sheet.cell(1, col).value)
        if header and header not in headers:
            headers[header] = col
    return headers


def safe_name(value: str) -> str:
    value = clean(value)
    value = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value[:120] or "UNKNOWN"


def date_folder_name(ship_date: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", clean(ship_date))
    if not match:
        return clean(ship_date)
    year, month, day = match.groups()
    return f"{int(year)}年{int(month)}月{int(day)}日"


def default_output_root(ship_date: str) -> Path:
    home = Path.home()
    return home / "Desktop" / "Codex Folder" / "DropShipZone" / "Codex发货记录" / date_folder_name(ship_date)


def read_dp_orders(csv_path: Path) -> dict[str, str]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"Order ID", "Item Sku"}
        missing_headers = required - set(reader.fieldnames or [])
        if missing_headers:
            raise ValueError(f"DP order CSV missing columns: {sorted(missing_headers)}")

        order_to_skus: dict[str, set[str]] = defaultdict(set)
        for row in reader:
            order_id = clean(row.get("Order ID"))
            sku = clean(row.get("Item Sku"))
            if not order_id or not sku:
                raise ValueError(f"DP order CSV has blank Order ID or Item Sku: {row}")
            order_to_skus[order_id].add(sku)

    multi_sku_orders = {order_id: sorted(skus) for order_id, skus in order_to_skus.items() if len(skus) > 1}
    if multi_sku_orders:
        raise ValueError(
            "Some DP orders contain multiple SKUs and cannot be matched to one label automatically: "
            + json.dumps(multi_sku_orders, ensure_ascii=False)
        )

    return {order_id: next(iter(skus)) for order_id, skus in order_to_skus.items()}


def read_yunexpress_shipments(xlsx_path: Path) -> dict[str, Shipment]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook["订单信息"] if "订单信息" in workbook.sheetnames else workbook.worksheets[0]
    headers = {clean(sheet.cell(1, col).value): col for col in range(1, sheet.max_column + 1)}

    required = ["客户单号", "运单号"]
    missing_headers = [header for header in required if header not in headers]
    if missing_headers:
        raise ValueError(f"YunExpress order file missing columns: {missing_headers}")

    tracking_col = headers.get("跟踪号")
    status_col = headers.get("订单状态")
    shipments: dict[str, Shipment] = {}
    duplicate_waybills: list[str] = []

    for row in range(2, sheet.max_row + 1):
        order_id = clean(sheet.cell(row, headers["客户单号"]).value)
        waybill = clean(sheet.cell(row, headers["运单号"]).value)
        tracking = clean(sheet.cell(row, tracking_col).value) if tracking_col else ""
        status = clean(sheet.cell(row, status_col).value) if status_col else ""
        if not order_id and not waybill:
            continue
        if not order_id or not waybill:
            raise ValueError(f"YunExpress row {row} has blank 客户单号 or 运单号")
        if waybill in shipments:
            duplicate_waybills.append(waybill)
        shipments[waybill] = Shipment(order_id=order_id, waybill=waybill, tracking=tracking, status=status)

    if duplicate_waybills:
        raise ValueError(f"Duplicate YunExpress waybills: {duplicate_waybills}")
    if not shipments:
        raise ValueError("No YunExpress shipment rows found.")
    return shipments


def read_sku_chinese_names(xlsx_path: Path) -> dict[str, str]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook.worksheets[0]
    headers = header_columns(sheet)
    sku_col = headers.get("SKU", 4)
    zh_col = headers.get("申报中文名", 2)

    sku_to_zh: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        sku = clean(sheet.cell(row, sku_col).value)
        zh = clean(sheet.cell(row, zh_col).value)
        if sku:
            sku_to_zh[sku] = zh
    return sku_to_zh


def iter_label_pdfs(zip_path: Path) -> Iterable[LabelPdf]:
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename
            if not name.lower().endswith(".pdf"):
                continue
            match = WAYBILL_RE.search(Path(name).stem)
            waybill = match.group(1).upper() if match else ""
            yield LabelPdf(zip_name=name, waybill=waybill)


def validate(
    labels: list[LabelPdf],
    shipments: dict[str, Shipment],
    order_to_sku: dict[str, str],
    sku_to_zh: dict[str, str],
) -> list[str]:
    errors: list[str] = []
    if not labels:
        errors.append("ZIP contains no PDF files.")

    blank_waybill_files = [label.zip_name for label in labels if not label.waybill]
    if blank_waybill_files:
        errors.append("PDF filenames without YT waybill: " + json.dumps(blank_waybill_files, ensure_ascii=False))

    pdf_waybill_counts = Counter(label.waybill for label in labels if label.waybill)
    duplicate_pdf_waybills = sorted([waybill for waybill, count in pdf_waybill_counts.items() if count > 1])
    if duplicate_pdf_waybills:
        errors.append("Duplicate PDF waybills in ZIP: " + json.dumps(duplicate_pdf_waybills, ensure_ascii=False))

    unmatched_pdf_waybills = sorted([waybill for waybill in pdf_waybill_counts if waybill not in shipments])
    if unmatched_pdf_waybills:
        errors.append("PDF waybills not found in YunExpress order file: " + json.dumps(unmatched_pdf_waybills, ensure_ascii=False))

    missing_pdf_waybills = sorted([waybill for waybill in shipments if waybill not in pdf_waybill_counts])
    if missing_pdf_waybills:
        errors.append("YunExpress waybills without matching PDF: " + json.dumps(missing_pdf_waybills, ensure_ascii=False))

    missing_dp_orders = []
    missing_sku_names = []
    for shipment in shipments.values():
        sku = order_to_sku.get(shipment.order_id)
        if not sku:
            missing_dp_orders.append(shipment.order_id)
            continue
        if not sku_to_zh.get(sku):
            missing_sku_names.append(sku)

    if missing_dp_orders:
        errors.append("YunExpress orders not found in DP order CSV: " + json.dumps(sorted(set(missing_dp_orders)), ensure_ascii=False))
    if missing_sku_names:
        errors.append("SKUs missing declaration Chinese name: " + json.dumps(sorted(set(missing_sku_names)), ensure_ascii=False))

    return errors


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_package(
    zip_path: Path,
    yunexpress_xlsx: Path,
    dp_orders_csv: Path,
    sku_xlsx: Path,
    ship_date: str,
    output_root: Path,
) -> Path:
    order_to_sku = read_dp_orders(dp_orders_csv)
    shipments = read_yunexpress_shipments(yunexpress_xlsx)
    sku_to_zh = read_sku_chinese_names(sku_xlsx)
    labels = list(iter_label_pdfs(zip_path))

    errors = validate(labels, shipments, order_to_sku, sku_to_zh)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = output_root / f"面单按SKU_{ship_date}_{timestamp}"
    run_dir.mkdir(parents=True, exist_ok=False)

    report_base = {
        "zip": str(zip_path),
        "yunexpress_xlsx": str(yunexpress_xlsx),
        "dp_orders_csv": str(dp_orders_csv),
        "sku_xlsx": str(sku_xlsx),
        "ship_date": ship_date,
        "pdf_count": len(labels),
        "yunexpress_shipment_count": len(shipments),
    }

    if errors:
        report = {**report_base, "status": "FAILED", "errors": errors}
        (run_dir / "校验失败报告.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        raise ValueError(f"Validation failed. Report written to: {run_dir / '校验失败报告.json'}")

    sku_counts: Counter[str] = Counter()
    for label in labels:
        shipment = shipments[label.waybill]
        sku_counts[order_to_sku[shipment.order_id]] += 1

    folder_name_by_sku = {
        sku: safe_name(f"{ship_date}_{sku}_{sku_to_zh[sku]}_{count}单")
        for sku, count in sku_counts.items()
    }

    detail_rows_by_sku: dict[str, list[dict[str, object]]] = defaultdict(list)
    copied_rows: list[dict[str, object]] = []

    with zipfile.ZipFile(zip_path) as zf:
        for label in labels:
            shipment = shipments[label.waybill]
            sku = order_to_sku[shipment.order_id]
            zh = sku_to_zh[sku]
            folder = run_dir / folder_name_by_sku[sku]
            folder.mkdir(parents=True, exist_ok=True)

            target_name = safe_name(f"{shipment.order_id}_{label.waybill}") + ".pdf"
            target_path = folder / target_name
            with zf.open(label.zip_name) as src, target_path.open("wb") as dst:
                shutil.copyfileobj(src, dst)

            row = {
                "日期": ship_date,
                "SKU": sku,
                "申报中文名": zh,
                "客户单号": shipment.order_id,
                "运单号": shipment.waybill,
                "跟踪号": shipment.tracking,
                "云途状态": shipment.status,
                "PDF文件": target_name,
            }
            detail_rows_by_sku[sku].append(row)
            copied_rows.append(row)

    for sku in detail_rows_by_sku:
        folder = run_dir / folder_name_by_sku[sku]
        zip_file = folder / f"{folder_name_by_sku[sku]}.zip"
        with zipfile.ZipFile(zip_file, "w", compression=zipfile.ZIP_DEFLATED) as supplier_zip:
            for pdf_path in sorted(folder.glob("*.pdf")):
                supplier_zip.write(pdf_path, arcname=pdf_path.name)

    detail_fields = ["日期", "SKU", "申报中文名", "客户单号", "运单号", "跟踪号", "云途状态", "PDF文件"]
    for sku, rows in detail_rows_by_sku.items():
        rows.sort(key=lambda row: str(row["客户单号"]))

    summary_rows = []
    for sku, rows in sorted(detail_rows_by_sku.items()):
        summary_rows.append({
            "日期": ship_date,
            "SKU": sku,
            "申报中文名": sku_to_zh[sku],
            "面单数量": len(rows),
            "订单数量": len({row["客户单号"] for row in rows}),
            "文件夹": folder_name_by_sku[sku],
        })
    write_csv(run_dir / "SKU分组汇总.csv", summary_rows, ["日期", "SKU", "申报中文名", "面单数量", "订单数量", "文件夹"])
    write_csv(run_dir / "全部面单明细.csv", sorted(copied_rows, key=lambda row: str(row["客户单号"])), detail_fields)

    report = {
        **report_base,
        "status": "OK",
        "output_dir": str(run_dir),
        "sku_count": len(summary_rows),
        "sku_summary": summary_rows,
    }
    (run_dir / "校验成功报告.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return run_dir


def main() -> int:
    parser = argparse.ArgumentParser(description="Sort YunExpress label PDFs by SKU.")
    parser.add_argument("--zip", required=True, type=Path, help="YunExpress label ZIP file.")
    parser.add_argument("--yunexpress-xlsx", required=True, type=Path, help="YunExpress order info xlsx.")
    parser.add_argument("--dp-orders-csv", required=True, type=Path, help="DP exported orders CSV.")
    parser.add_argument("--sku-xlsx", required=True, type=Path, help="SKU declaration workbook.")
    parser.add_argument("--date", required=True, help="Date prefix for folders, e.g. 2026-07-09.")
    parser.add_argument(
        "--output-root",
        type=Path,
        default=None,
        help="Output root directory. Defaults to Codex发货记录/<date folder>.",
    )
    args = parser.parse_args()
    output_root = args.output_root or default_output_root(args.date)

    try:
        run_dir = build_package(
            zip_path=args.zip,
            yunexpress_xlsx=args.yunexpress_xlsx,
            dp_orders_csv=args.dp_orders_csv,
            sku_xlsx=args.sku_xlsx,
            ship_date=args.date,
            output_root=output_root,
        )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"OK: {run_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

