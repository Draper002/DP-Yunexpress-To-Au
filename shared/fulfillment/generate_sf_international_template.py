#!/usr/bin/env python3
"""Generate an SF International batch import workbook from DP AU orders."""

from __future__ import annotations

import argparse
import json
import posixpath
import sys
from collections import Counter
from copy import deepcopy
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.utils import get_column_letter

from generate_yunexpress_template import (
    clean,
    default_output_root,
    parse_number,
    read_dp_orders,
    read_sku_info,
    round3,
    unique_path,
)


DEFAULT_COUNTRY = "AU"
DEFAULT_CURRENCY = "USD/美元"
DEFAULT_STATUS = "processing"

SF_BUSINESS_TYPES = (
    "国际小包",
    "国际电商专递CD（普货）",
    "国际电商专递-标准",
    "国际电商专递-快速",
    "国际电商专递-优先CD",
    "敏货电商专递",
    "服装专线",
    "轻抛专线",
    "9610专线",
    "国际经济小包挂号",
    "轻小件小包挂号",
    "国际电商专递-丰选",
    "国际电商专递-商派专线",
    "商派服装专线",
    "特货电商专递",
    "国际电商专递-优选",
)

BATTERY_OPTIONS = ("否", "是")

HEADER = {
    "business_type": "业务类型*[0]",
    "order_id": "用户订单号*[1]",
    "recipient": "收件联系人*[14]",
    "country": "目的地国家代码*[15]",
    "state": "收方州/省[16]",
    "city": "收方城市*[17]",
    "address": "收方详细地址*[20]",
    "postcode": "收件邮编*[22]",
    "phone": "收件人电话*[23]",
    "mobile": "收件人手机*[24]",
    "total_weight": "总重量(KG)[31]",
    "battery": "是否带电*[38]",
    "currency": "申报价值币种*",
    "packing_info": "配货信息[39]",
    "english_name": "英文申报品名1*[42]",
    "chinese_name": "中文申报品名1*[43]",
    "unit_price": "申报价值（单价）1*[44]",
    "unit_weight": "品名1单位重量*[46]",
    "quantity": "申报品数量1*[47]",
    "merchant_sku": "品名1商家产品识别码[96]",
}

REQUIRED_KEYS = (
    "business_type",
    "order_id",
    "recipient",
    "country",
    "city",
    "address",
    "postcode",
    "phone",
    "mobile",
    "battery",
    "currency",
    "english_name",
    "chinese_name",
    "unit_price",
    "unit_weight",
    "quantity",
)


def header_columns(sheet) -> dict[str, int]:
    return {
        clean(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
        if clean(sheet.cell(1, column).value)
    }


def validate_settings(business_type: str, battery: str) -> None:
    if business_type not in SF_BUSINESS_TYPES:
        raise ValueError(
            "Invalid SF business type. Choose one value from the template list: "
            + json.dumps(SF_BUSINESS_TYPES, ensure_ascii=False)
        )
    if battery not in BATTERY_OPTIONS:
        raise ValueError("SF battery option must be 是 or 否.")


def build_sf_rows(
    orders: list[dict[str, str]],
    sku_map: dict[str, dict[str, object]],
    business_type: str,
    battery: str,
) -> list[dict[str, object]]:
    validate_settings(business_type, battery)
    output_rows: list[dict[str, object]] = []
    missing_skus: list[str] = []

    for order in orders:
        sku = clean(order.get("Item Sku"))
        info = sku_map.get(sku)
        if (
            not info
            or not info.get("chinese_name")
            or not info.get("english_name")
            or info.get("price_usd") is None
            or info.get("unit_weight_kg") is None
        ):
            missing_skus.append(sku)
            continue

        quantity = parse_number(order.get("Item Quantity"))
        if quantity is None or quantity <= 0:
            raise ValueError(
                f"Invalid quantity for Order ID {clean(order.get('Order ID'))}: {order.get('Item Quantity')}"
            )

        unit_weight = float(info["unit_weight_kg"])
        output_rows.append(
            {
                "business_type": business_type,
                "order_id": clean(order.get("Order ID")),
                "recipient": clean(order.get("Ship to Name")),
                "country": DEFAULT_COUNTRY,
                "state": clean(order.get("State")),
                "city": clean(order.get("Suburb")),
                "address": clean(order.get("Street")),
                "postcode": clean(order.get("Postcode")),
                "phone": clean(order.get("Phone Number")),
                "mobile": clean(order.get("Phone Number")),
                "total_weight": round3(unit_weight * quantity),
                "battery": battery,
                "currency": DEFAULT_CURRENCY,
                "packing_info": sku,
                "english_name": info["english_name"],
                "chinese_name": info["chinese_name"],
                "unit_price": float(info["price_usd"]),
                "unit_weight": unit_weight,
                "quantity": int(quantity) if quantity.is_integer() else quantity,
                "merchant_sku": sku,
            }
        )

    if missing_skus:
        raise ValueError(
            "Missing or incomplete SKU declaration data: "
            + json.dumps(sorted(set(missing_skus)), ensure_ascii=False)
        )
    return output_rows


SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def worksheet_part_name(archive: ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    sheets = workbook_root.find(f"{{{SHEET_NS}}}sheets")
    if sheets is None or not list(sheets):
        raise ValueError("SF template has no worksheet.")
    sheet = next(
        (item for item in sheets if item.get("name") == sheet_name),
        list(sheets)[0],
    )
    relation_id = sheet.get(f"{{{REL_NS}}}id")
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relation = next(
        (item for item in relationships if item.get("Id") == relation_id),
        None,
    )
    if relation is None:
        raise ValueError("SF template worksheet relationship is missing.")
    target = relation.get("Target", "")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def clear_cell_value(cell: ET.Element) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)


def set_cell_value(cell: ET.Element, value: object) -> None:
    clear_cell_value(cell)
    if value is None:
        return
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.set("t", "n")
        element = ET.SubElement(cell, f"{{{SHEET_NS}}}v")
        element.text = str(value)
        return
    cell.set("t", "inlineStr")
    inline_string = ET.SubElement(cell, f"{{{SHEET_NS}}}is")
    text = ET.SubElement(inline_string, f"{{{SHEET_NS}}}t")
    text.text = str(value)


def write_rows_to_template_xml(
    template_path: Path,
    output_path: Path,
    sheet_name: str,
    rows: list[dict[str, object]],
    columns: dict[str, int],
) -> None:
    ET.register_namespace("", SHEET_NS)
    ET.register_namespace("r", REL_NS)

    temporary_path = output_path.with_suffix(output_path.suffix + ".tmp")
    with ZipFile(template_path, "r") as source:
        sheet_part = worksheet_part_name(source, sheet_name)
        sheet_root = ET.fromstring(source.read(sheet_part))
        sheet_data = sheet_root.find(f"{{{SHEET_NS}}}sheetData")
        if sheet_data is None:
            raise ValueError("SF template worksheet has no sheetData section.")

        existing_rows = list(sheet_data.findall(f"{{{SHEET_NS}}}row"))
        template_row = next((row for row in existing_rows if row.get("r") == "3"), None)
        if template_row is None:
            template_row = next((row for row in existing_rows if row.get("r") == "2"), None)
        if template_row is None:
            raise ValueError("SF template has no reusable data row formatting.")

        for row in existing_rows:
            if int(row.get("r", "0")) >= 2:
                sheet_data.remove(row)

        reserved_rows = max(9, len(rows))
        for offset in range(reserved_rows):
            row_number = offset + 2
            row_element = deepcopy(template_row)
            row_element.set("r", str(row_number))
            cell_by_column: dict[int, ET.Element] = {}
            for cell in row_element.findall(f"{{{SHEET_NS}}}c"):
                reference = cell.get("r", "A3")
                column_letters = "".join(char for char in reference if char.isalpha())
                column_number = openpyxl.utils.column_index_from_string(column_letters)
                cell.set("r", f"{column_letters}{row_number}")
                clear_cell_value(cell)
                cell_by_column[column_number] = cell

            values = rows[offset] if offset < len(rows) else {}
            for key, value in values.items():
                column_number = columns[HEADER[key]]
                cell = cell_by_column.get(column_number)
                if cell is None:
                    cell = ET.Element(
                        f"{{{SHEET_NS}}}c",
                        {"r": f"{get_column_letter(column_number)}{row_number}"},
                    )
                    row_element.append(cell)
                set_cell_value(cell, value)
            sheet_data.append(row_element)

        dimension = sheet_root.find(f"{{{SHEET_NS}}}dimension")
        if dimension is not None:
            dimension.set("ref", f"A1:CX{reserved_rows + 1}")
        sheet_xml = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)

        with ZipFile(temporary_path, "w", compression=ZIP_DEFLATED) as target:
            for item in source.infolist():
                if item.filename == sheet_part:
                    target.writestr(item, sheet_xml)
                else:
                    target.writestr(item, source.read(item.filename))

    temporary_path.replace(output_path)


def write_template(
    template_path: Path,
    rows: list[dict[str, object]],
    ship_date: str,
    business_type: str,
    output_root: Path,
) -> Path:
    workbook = openpyxl.load_workbook(template_path, data_only=False, read_only=True)
    try:
        sheet = workbook["标准上传格式"] if "标准上传格式" in workbook.sheetnames else workbook.worksheets[0]
        columns = header_columns(sheet)
        sheet_name = sheet.title

        missing_headers = sorted(header for header in HEADER.values() if header not in columns)
        if missing_headers:
            raise ValueError("SF template missing headers: " + json.dumps(missing_headers, ensure_ascii=False))
    finally:
        workbook.close()

    output_dir = output_root / "顺丰国际上传模板"
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_business_type = business_type.replace("/", "-").replace("\\", "-")
    output_path = unique_path(
        output_dir / f"顺丰国际批量寄件_{ship_date}_{len(rows)}orders_{safe_business_type}.xlsm"
    )
    write_rows_to_template_xml(template_path, output_path, sheet_name, rows, columns)
    return output_path


def validate_output(path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    try:
        sheet = workbook["标准上传格式"] if "标准上传格式" in workbook.sheetnames else workbook.worksheets[0]
        columns = header_columns(sheet)
        order_id_col = columns[HEADER["order_id"]]
        row_indexes = [
            row for row in range(2, sheet.max_row + 1) if clean(sheet.cell(row, order_id_col).value)
        ]
        missing = [
            (row, key)
            for row in row_indexes
            for key in REQUIRED_KEYS
            if sheet.cell(row, columns[HEADER[key]]).value in (None, "")
        ]
        order_ids = [clean(sheet.cell(row, order_id_col).value) for row in row_indexes]
        duplicate_order_ids = sorted(
            order_id for order_id, count in Counter(order_ids).items() if count > 1
        )
        return {
            "row_count": len(row_indexes),
            "required_missing_count": len(missing),
            "duplicate_order_ids": duplicate_order_ids,
            "business_types": dict(
                Counter(clean(sheet.cell(row, columns[HEADER["business_type"]]).value) for row in row_indexes)
            ),
            "countries": dict(
                Counter(clean(sheet.cell(row, columns[HEADER["country"]]).value) for row in row_indexes)
            ),
            "currencies": dict(
                Counter(clean(sheet.cell(row, columns[HEADER["currency"]]).value) for row in row_indexes)
            ),
            "battery": dict(
                Counter(clean(sheet.cell(row, columns[HEADER["battery"]]).value) for row in row_indexes)
            ),
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate SF International batch template from DP orders.")
    parser.add_argument("--dp-orders-csv", required=True, type=Path)
    parser.add_argument("--sku-xlsx", required=True, type=Path)
    parser.add_argument("--sf-template-xlsm", required=True, type=Path)
    parser.add_argument("--business-type", required=True)
    parser.add_argument("--battery", choices=BATTERY_OPTIONS, default="否")
    parser.add_argument("--date", required=True)
    parser.add_argument("--status", default=DEFAULT_STATUS)
    parser.add_argument("--output-root", type=Path, default=None)
    args = parser.parse_args()

    try:
        output_root = args.output_root or default_output_root(args.date)
        orders, source_counts = read_dp_orders(args.dp_orders_csv, args.status)
        sku_map = read_sku_info(args.sku_xlsx)
        output_rows = build_sf_rows(orders, sku_map, args.business_type, args.battery)
        output_path = write_template(
            args.sf_template_xlsm,
            output_rows,
            args.date,
            args.business_type,
            output_root,
        )
        validation = validate_output(output_path)
        if validation["required_missing_count"] != 0 or validation["duplicate_order_ids"]:
            raise ValueError(f"Generated SF file failed validation: {validation}")
        report = {
            "status": "OK",
            "source_counts": source_counts,
            "output_path": str(output_path),
            "validation": validation,
        }
        output_path.with_suffix(".report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"OK: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
