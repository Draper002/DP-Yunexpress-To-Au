import csv
import importlib
import os
import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient


class WebPlatformTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        os.environ["DP_WEB_DATA_DIR"] = cls.temp_dir.name
        os.environ["DP_WEB_ADMIN_EMAIL"] = "admin@test.local"
        os.environ["DP_WEB_ADMIN_PASSWORD"] = "test-password"
        os.environ["DP_WEB_COOKIE_SECURE"] = "0"

        cls.web = importlib.import_module("web.app")
        cls.client_context = TestClient(cls.web.app)
        cls.client = cls.client_context.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.client_context.__exit__(None, None, None)
        cls.temp_dir.cleanup()

    def login(self, email="admin@test.local", password="test-password"):
        self.client.cookies.clear()
        response = self.client.post(
            "/login",
            data={"email": email, "password": password},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        return response

    def test_home_exposes_both_platforms_for_all_three_steps(self):
        response = self.login()
        self.assertEqual(response.text.count('name="platform" value="sf"'), 3)
        self.assertIn("顺丰国际订单数据", response.text)
        self.assertIn("顺丰国际面单 PDF", response.text)
        self.assertIn("SF INTERNATIONAL", response.text)
        self.assertIn('name="merge_same_recipient_sku"', response.text)
        self.assertIn("合并相同收件人 + 相同 SKU", response.text)
        self.assertIn("getTimezoneOffset", response.text)

    def test_config_update_is_all_or_nothing(self):
        self.login()
        current_sku = self.web.config_path("sku")
        current_sku.parent.mkdir(parents=True, exist_ok=True)
        current_sku.write_bytes(b"existing-config")

        upload_path = Path(self.temp_dir.name) / "valid-sku.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["序号", "申报中文名", "申报英文名", "SKU", "单价", "单重"])
        sheet.append([1, "测试商品", "test product", "SKU-001", 10, 0.5])
        workbook.save(upload_path)
        workbook.close()

        response = self.client.post(
            "/config",
            files={
                "sku": (
                    "sku.xlsx",
                    upload_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                "yun_template": ("wrong.txt", b"not-an-xlsx", "text/plain"),
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(current_sku.read_bytes(), b"existing-config")
        self.assertEqual(list(current_sku.parent.glob(".staging-*")), [])

    def test_step2_rejects_a_different_order_batch_without_replacing_working_file(self):
        self.login()
        template_path = self.web.config_path("dp_template")
        template_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = openpyxl.Workbook()
        workbook.active.append(["Order ID", "Carrier", "Tracking Number"])
        workbook.save(template_path)
        workbook.close()

        batch = self.web.batch_dir(1, "yunexpress")
        shutil.rmtree(batch, ignore_errors=True)
        batch.mkdir(parents=True, exist_ok=True)
        with (batch / "dp_orders.csv").open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=["Order ID", "Status"])
            writer.writeheader()
            writer.writerow({"Order ID": "CURRENT-001", "Status": "processing"})
        previous = batch / "yun_orders.xlsx"
        previous.write_bytes(b"previous-valid-batch")

        upload_path = Path(self.temp_dir.name) / "wrong-yun-orders.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "订单信息"
        sheet.append(["客户单号", "运单号"])
        sheet.append(["OTHER-001", "YT000000000001"])
        workbook.save(upload_path)
        workbook.close()

        response = self.client.post(
            "/step/2",
            data={"date": "2026-07-26", "platform": "yunexpress"},
            files={
                "file": (
                    upload_path.name,
                    upload_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("does not match DP CSV", response.text)
        self.assertEqual(previous.read_bytes(), b"previous-valid-batch")

    def test_sf_step3_requires_the_step2_order_batch(self):
        self.login()
        batch = self.web.batch_dir(1, "sf")
        shutil.rmtree(batch, ignore_errors=True)

        response = self.client.post(
            "/step/3",
            data={"date": "2026-07-26", "platform": "sf"},
            files={"file": ("labels.pdf", b"%PDF-placeholder", "application/pdf")},
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("需要先在同一账号完成第 2 步", response.text)

    def test_platform_specific_config_requirements(self):
        self.assertEqual(
            self.web.required_config_keys(1, "sf"),
            ("sku", "sf_template"),
        )
        self.assertEqual(
            self.web.required_config_keys(2, "sf"),
            ("dp_template",),
        )
        self.assertEqual(self.web.required_config_keys(3, "sf"), ())
        self.assertEqual(
            self.web.required_config_keys(3, "yunexpress"),
            ("sku",),
        )

    def test_health_endpoint_lists_both_platforms(self):
        response = self.client.get("/healthz")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["platforms"], ["yunexpress", "sf"])
        self.assertIn("same_recipient_same_sku_merge", response.json()["features"])

    def test_employee_cannot_download_another_users_result(self):
        owner_email = "owner@test.local"
        other_email = "other@test.local"
        with self.web.connection() as conn:
            for email in (owner_email, other_email):
                conn.execute(
                    "INSERT OR IGNORE INTO users(email,password_hash,role,created_at) VALUES(?,?,?,?)",
                    (email, self.web.hash_password("employee-pass"), "employee", "2026-07-26T20:00:00"),
                )
            owner_id = conn.execute("SELECT id FROM users WHERE email=?", (owner_email,)).fetchone()["id"]
            conn.commit()

        result = Path(self.temp_dir.name) / "private-result.xlsx"
        result.write_bytes(b"private")
        self.web.add_run(owner_id, "步骤 2 · 顺丰国际", "2026-07-26", "SUCCESS", result)
        with self.web.connection() as conn:
            run_id = conn.execute("SELECT max(id) AS id FROM runs").fetchone()["id"]

        self.login(other_email, "employee-pass")
        response = self.client.get(f"/download/{run_id}", follow_redirects=False)
        self.assertEqual(response.status_code, 404)


if __name__ == "__main__":
    unittest.main()
