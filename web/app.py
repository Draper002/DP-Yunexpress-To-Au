from __future__ import annotations

import asyncio
import hashlib
import hmac
import html
import io
import os
import secrets
import shutil
import sqlite3
import sys
import threading
import time
from contextlib import asynccontextmanager, closing, redirect_stderr, redirect_stdout
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from starlette.concurrency import run_in_threadpool

ROOT = Path(__file__).resolve().parents[1]
TOOL_ROOT = ROOT / "shared" / "fulfillment"
DATA = Path(os.getenv("DP_WEB_DATA_DIR", str(Path(__file__).parent / "data"))).resolve()
CONFIG = DATA / "config"
RESULTS = DATA / "results"
WORKING = DATA / "working"
DB_PATH = DATA / "app.sqlite3"
ADMIN_EMAIL = os.getenv("DP_WEB_ADMIN_EMAIL", "admin@example.com")
ADMIN_PASSWORD = os.getenv("DP_WEB_ADMIN_PASSWORD", "change-me")
MAX_UPLOAD = 80 * 1024 * 1024
COOKIE_SECURE = os.getenv("DP_WEB_COOKIE_SECURE", "0") == "1"
UNSAFE_ADMIN_PASSWORDS = {"change-me", "replace-before-starting"}
sys.path.insert(0, str(TOOL_ROOT))
import generate_dp_shipment_upload
import generate_sf_international_template
import generate_yunexpress_template
import sort_yunexpress_labels_by_sku
import split_sf_labels_by_sender

CONFIG_FILES = {
    "sku": "sku.xlsx",
    "yun_template": "yun_template.xlsx",
    "sf_template": "sf_template.xlsm",
    "dp_template": "dp_template.xlsx",
}
CONFIG_LABELS = {
    "sku": "SKU å•†å“åº“",
    "yun_template": "äº‘é€”æ ‡å‡†æ¨¡æ¿",
    "sf_template": "é¡ºä¸°å›½é™…æ ‡å‡†æ¨¡æ¿",
    "dp_template": "DP å‘è´§æ¨¡æ¿",
}
PLATFORM_LABELS = {"yunexpress": "äº‘é€”", "sf": "é¡ºä¸°å›½é™…"}
SCRIPT_LOCK = threading.Lock()
WORKFLOW_LOCK = asyncio.Lock()
YUN_TEMPLATE_HEADERS = {
    "å®¢æˆ·è®¢å•å·",
    "äº§å“ä»£ç ",
    "æ”¶ä»¶äººå›½å®¶",
    "æ”¶ä»¶äººå§“å",
    "æ”¶ä»¶äººåœ°å€",
    "æ”¶ä»¶äººåŸŽå¸‚",
    "æ”¶ä»¶äººçœå·ž",
    "é‚®ç¼–",
    "ç”µè¯",
    "ä»¶æ•°",
    "åŒ…è£¹æ€»é‡é‡",
    "ç”³æŠ¥å¸ç§",
    "SKU1",
    "ç”³æŠ¥å“å1",
    "ä¸­æ–‡ç”³æŠ¥å“å1",
    "ç”³æŠ¥æ•°é‡1",
    "ç”³æŠ¥FOBä»·1",
    "ç”³æŠ¥å•é‡1",
}


@asynccontextmanager
async def lifespan(_app: FastAPI):
    if COOKIE_SECURE and ADMIN_PASSWORD in UNSAFE_ADMIN_PASSWORDS:
        raise RuntimeError("Set a private DP_WEB_ADMIN_PASSWORD before secure deployment.")
    init_db()
    for folder in (CONFIG, RESULTS, WORKING):
        folder.mkdir(parents=True, exist_ok=True)
    yield


app = FastAPI(title="DP International Fulfillment To Au Web", lifespan=lifespan)


def connection():
    DATA.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    value = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 240_000)
    return f"{salt.hex()}${value.hex()}"


def check_password(password: str, encoded: str) -> bool:
    try:
        salt, expected = encoded.split("$", 1)
        actual = hash_password(password, bytes.fromhex(salt)).split("$", 1)[1]
        return hmac.compare_digest(actual, expected)
    except (ValueError, TypeError):
        return False


def init_db():
    DATA.mkdir(parents=True, exist_ok=True)
    with closing(connection()) as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY,email TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL,active INTEGER NOT NULL DEFAULT 1,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS sessions(token TEXT PRIMARY KEY,user_id INTEGER NOT NULL,expires_at INTEGER NOT NULL);
        CREATE TABLE IF NOT EXISTS runs(id INTEGER PRIMARY KEY,user_id INTEGER NOT NULL,step TEXT NOT NULL,status TEXT NOT NULL,date TEXT NOT NULL,result_path TEXT,report_path TEXT,created_at TEXT NOT NULL,error TEXT);
        """)
        if not conn.execute("SELECT 1 FROM users LIMIT 1").fetchone():
            conn.execute("INSERT INTO users(email,password_hash,role,created_at) VALUES(?,?,?,?)", (ADMIN_EMAIL, hash_password(ADMIN_PASSWORD), "admin", datetime.now().isoformat(timespec="seconds")))
        conn.execute("DELETE FROM sessions WHERE expires_at<=?", (int(time.time()),))
        conn.commit()


def user_for(request: Request):
    token = request.cookies.get("dp_session")
    if not token:
        return None
    with closing(connection()) as conn:
        return conn.execute("SELECT users.* FROM sessions JOIN users ON users.id=sessions.user_id WHERE token=? AND expires_at>? AND active=1", (token, int(time.time()))).fetchone()


async def save_upload(upload: UploadFile, folder: Path, extensions: set[str]) -> Path:
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix not in extensions:
        raise ValueError("æ–‡ä»¶ç±»åž‹ä¸æ­£ç¡®")
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / f"{secrets.token_hex(10)}{suffix}"
    size = 0
    with target.open("wb") as out:
        while chunk := await upload.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_UPLOAD:
                target.unlink(missing_ok=True)
                raise ValueError("æ–‡ä»¶è¶…è¿‡ 80 MB é™åˆ¶")
            out.write(chunk)
    return target


def config_path(key: str) -> Path:
    if key not in CONFIG_FILES:
        raise KeyError(f"Unknown config key: {key}")
    return CONFIG / CONFIG_FILES[key]


def validate_config_upload(key: str, path: Path) -> None:
    if key == "sku":
        if not generate_yunexpress_template.read_sku_info(path):
            raise ValueError("SKU å•†å“åº“æ²¡æœ‰å¯ç”¨çš„ SKU æ•°æ®")
        return

    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=key == "sf_template",
    )
    try:
        if not workbook.worksheets:
            raise ValueError(f"{CONFIG_LABELS[key]}æ²¡æœ‰å·¥ä½œè¡¨")
        sheet = (
            workbook["æ ‡å‡†ä¸Šä¼ æ ¼å¼"]
            if key == "sf_template" and "æ ‡å‡†ä¸Šä¼ æ ¼å¼" in workbook.sheetnames
            else workbook.worksheets[0]
        )
        headers = {
            generate_yunexpress_template.clean(sheet.cell(1, column).value)
            for column in range(1, sheet.max_column + 1)
        }
        if key == "yun_template":
            missing = sorted(YUN_TEMPLATE_HEADERS - headers)
            if missing:
                raise ValueError("äº‘é€”æ ‡å‡†æ¨¡æ¿ç¼ºå°‘å­—æ®µï¼š" + "ã€".join(missing))
        elif key == "sf_template":
            missing = sorted(set(generate_sf_international_template.HEADER.values()) - headers)
            if missing:
                raise ValueError("é¡ºä¸°å›½é™…æ ‡å‡†æ¨¡æ¿ç¼ºå°‘å­—æ®µï¼š" + "ã€".join(missing))
    finally:
        workbook.close()


def normalize_platform(value: str) -> str:
    platform = (value or "").strip().lower()
    if platform not in PLATFORM_LABELS:
        raise ValueError("è¯·é€‰æ‹©äº‘é€”æˆ–é¡ºä¸°å›½é™…")
    return platform


def normalize_date(value: str) -> str:
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("å‘è´§æ—¥æœŸæ ¼å¼å¿…é¡»ä¸º YYYY-MM-DD") from exc
    if parsed.strftime("%Y-%m-%d") != value:
        raise ValueError("å‘è´§æ—¥æœŸæ ¼å¼å¿…é¡»ä¸º YYYY-MM-DD")
    return value


def required_config_keys(step: int, platform: str) -> tuple[str, ...]:
    platform = normalize_platform(platform)
    requirements = {
        (1, "yunexpress"): ("sku", "yun_template"),
        (1, "sf"): ("sku", "sf_template"),
        (2, "yunexpress"): ("dp_template",),
        (2, "sf"): ("dp_template",),
        (3, "yunexpress"): ("sku",),
        (3, "sf"): (),
    }
    try:
        return requirements[(step, platform)]
    except KeyError as exc:
        raise ValueError("æœªçŸ¥å¤„ç†æ­¥éª¤") from exc


def add_run(user_id: int, step: str, date: str, status: str, result: Path | None = None, report: Path | None = None, error: str = ""):
    with closing(connection()) as conn:
        conn.execute("INSERT INTO runs(user_id,step,status,date,result_path,report_path,created_at,error) VALUES(?,?,?,?,?,?,?,?)", (user_id, step, status, date, str(result) if result else None, str(report) if report else None, datetime.now().isoformat(timespec="seconds"), error[:3000]))
        conn.commit()


def script(module, args: list[str]) -> tuple[int, str, str]:
    with SCRIPT_LOCK:
        old = sys.argv[:]
        stdout, stderr = io.StringIO(), io.StringIO()
        try:
            sys.argv = [module.__name__] + args
            with redirect_stdout(stdout), redirect_stderr(stderr):
                code = module.main()
            return int(code or 0), stdout.getvalue(), stderr.getvalue()
        finally:
            sys.argv = old


def result_from(output: str) -> Path:
    paths = [line[3:].strip().strip('"') for line in output.splitlines() if line.startswith("OK:")]
    if not paths or not Path(paths[-1]).exists():
        raise ValueError(output.strip() or "å¤„ç†è„šæœ¬æ²¡æœ‰è¿”å›žç»“æžœ")
    return Path(paths[-1])


def batch_dir(user_id: int, platform: str) -> Path:
    path = WORKING / str(user_id) / normalize_platform(platform)
    path.mkdir(parents=True, exist_ok=True)
    return path


def field(name: str, label: str, accept: str, required: bool = True) -> str:
    required_attr = " required" if required else ""
    return f'<label class="file"><b>{label}</b><input name="{name}" type="file" accept="{accept}"{required_attr}><span>é€‰æ‹©æ–‡ä»¶</span></label>'


def platform_selector() -> str:
    return '''<div class="platform-picker" role="group" aria-label="æ‰¿è¿å¹³å°">
    <label><input type="radio" name="platform" value="yunexpress" checked><span>äº‘é€”</span></label>
    <label><input type="radio" name="platform" value="sf"><span>é¡ºä¸°å›½é™…</span></label>
    </div>'''


def sf_options() -> str:
    business_options = "".join(
        f'<option value="{html.escape(value)}">{html.escape(value)}</option>'
        for value in generate_sf_international_template.SF_BUSINESS_TYPES
    )
    battery_options = "".join(
        f'<label><input type="radio" name="battery" value="{html.escape(value)}"{(" checked" if value == "å¦" else "")}><span>{html.escape(value)}</span></label>'
        for value in generate_sf_international_template.BATTERY_OPTIONS
    )
    return f'''<div class="sf-only sf-options">
    <label>ä¸šåŠ¡ç±»åž‹<select name="business_type"><option value="">è¯·é€‰æ‹©ä¸šåŠ¡ç±»åž‹</option>{business_options}</select></label>
    <div><b>æ˜¯å¦å¸¦ç”µ</b><div class="mini-segment">{battery_options}</div></div>
    </div>'''


def page(title: str, content: str, user=None) -> str:
    nav = f'<header><a class="brand" href="/">DP å›½é™…å‘è´§ <small>TO AU</small></a><nav><a href="/">å·¥ä½œå°</a>{"<a href=/admin>æˆå‘˜ç®¡ç†</a>" if user and user["role"] == "admin" else ""}<a href="/logout">é€€å‡º</a></nav></header>' if user else '<header><a class="brand" href="/login">DP å›½é™…å‘è´§ <small>TO AU</small></a></header>'
    return f'<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{html.escape(title)} Â· DP International Fulfillment</title><style>{CSS}</style></head><body>{nav}<main>{content}</main><script>{SCRIPT}</script></body></html>'


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if user_for(request):
        return RedirectResponse("/", status_code=303)
    return HTMLResponse(page("ç™»å½•", '<section class="login"><div class="mark">DP<span>&</span></div><p class="eyebrow">PRIVATE OPERATIONS</p><h1>ç™»å½•å‘è´§å·¥ä½œå°</h1><p class="lead">å‘˜å·¥è´¦å·ç”±ç®¡ç†å‘˜åˆ›å»ºï¼Œè®¢å•æ–‡ä»¶å’Œå¤„ç†è®°å½•æŒ‰è´¦å·éš”ç¦»ä¿å­˜ã€‚</p><form method="post" action="/login"><label>é‚®ç®±<input name="email" type="email" required></label><label>å¯†ç <input name="password" type="password" required></label><button class="primary" type="submit">ç™»å½•</button></form></section>'))


@app.post("/login")
def login(email: str = Form(...), password: str = Form(...)):
    with closing(connection()) as conn:
        row = conn.execute("SELECT * FROM users WHERE lower(email)=lower(?) AND active=1", (email.strip(),)).fetchone()
    if not row or not check_password(password, row["password_hash"]):
        return HTMLResponse(page("ç™»å½•å¤±è´¥", '<section class="login"><h1>æ— æ³•ç™»å½•</h1><p class="lead">é‚®ç®±æˆ–å¯†ç ä¸æ­£ç¡®ã€‚</p><a class="button" href="/login">è¿”å›žç™»å½•</a></section>'), status_code=401)
    token = secrets.token_urlsafe(32)
    with closing(connection()) as conn:
        conn.execute("INSERT INTO sessions VALUES(?,?,?)", (token, row["id"], int(time.time()) + 86400)); conn.commit()
    response = RedirectResponse("/", status_code=303)
    response.set_cookie("dp_session", token, httponly=True, samesite="lax", secure=COOKIE_SECURE, max_age=86400)
    return response


@app.get("/logout")
def logout(request: Request):
    token = request.cookies.get("dp_session")
    if token:
        with closing(connection()) as conn:
            conn.execute("DELETE FROM sessions WHERE token=?", (token,)); conn.commit()
    response = RedirectResponse("/login", status_code=303); response.delete_cookie("dp_session"); return response


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    user = user_for(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    with closing(connection()) as conn:
        runs = conn.execute(
            "SELECT * FROM runs WHERE user_id=? ORDER BY id DESC LIMIT 10",
            (user["id"],),
        ).fetchall()
    records = ""
    for record in runs:
        result_link = f'<a href="/download/{record["id"]}">ä¸‹è½½ç»“æžœ</a>' if record["result_path"] and Path(record["result_path"]).exists() else "-"
        records += f'<tr><td>{html.escape(record["created_at"])}</td><td>{html.escape(record["step"])}</td><td><span class="status {record["status"].lower()}">{html.escape(record["status"])}</span></td><td>{html.escape(record["date"])}</td><td>{result_link}</td></tr>'
    records = records or '<tr><td colspan="5" class="empty">è¿˜æ²¡æœ‰å¤„ç†è®°å½•ï¼Œä»Žç¬¬ 1 æ­¥å¼€å§‹ã€‚</td></tr>'
    today = datetime.now().date().isoformat()
    config_ready = all(config_path(key).exists() for key in CONFIG_LABELS)
    content = f'''<section class="hero"><div><p class="eyebrow">AU FULFILLMENT DESK</p><h1>å‘è´§å·¥ä½œå°</h1><p class="lead">é€‰æ‹©äº‘é€”æˆ–é¡ºä¸°å›½é™…ï¼ŒæŒ‰é¡ºåºå®Œæˆå¯„ä»¶æ¨¡æ¿ã€DP è¿å•å·å›žå¡«å’Œä¾›åº”å•†é¢å•æ•´ç†ã€‚</p></div><div class="identity"><i></i><div><b>{html.escape(user["email"])}</b><small>{"ç®¡ç†å‘˜" if user["role"] == "admin" else "å‘˜å·¥"} Â· åœ¨çº¿</small></div></div></section>
    <section class="config"><div><small>å›ºå®šé…ç½®</small><b>{"å…¨éƒ¨å°±ç»ª" if config_ready else "éœ€è¦æ£€æŸ¥"}</b><em>SßŸ9¶‰žËkºwµçC–æÏ–>Ã–º3š"Cž²°€Äƒš¶”ˆ¤(€€€€€€€¥˜Á±…Ñ™½É´€ôô€‰åÕ¹•áÁÉ•ÍÌˆè(€€€€€€€€€€€½µµ¥Ñ}Ñ…É•Ð€ô‰…Ñ €¼€‰åÕ¹}½É‘•ÉÌ¹á±Íàˆ(€€€€€€€€€€€Í½ÕÉ•}…ÉÌ€ôlˆ´µåÕ¹•áÁÉ•ÍÌµá±Íàˆ°ÍÑÈ¡Í½ÕÉ”¥t(€€€€€€€•±Í”è(€€€€€€€€€€€½µµ¥Ñ}Ñ…É•Ð€ô‰…Ñ €¼˜‰Í™}½É‘•ÉÍíÍ½ÕÉ”¹ÍÕ™™¥à¹±½Ý•È ¥ôˆ(€€€€€€€€€€€Í½ÕÉ•}…ÉÌ€ôlˆ´µÍ˜µ¥¹Ñ•É¹…Ñ¥½¹…°µ™¥±”ˆ°ÍÑÈ¡Í½ÕÉ”¥t(€€€€€€€½‘”°½ÕÐ°•ÉÈ€ôÍÉ¥ÁÐ (€€€€€€€€€€€•¹•É…Ñ•}‘Á}Í¡¥Áµ•¹Ñ}ÕÁ±½…°(€€€€€€€€€€€Í½ÕÉ•}…ÉÌ(€€€€€€€€€€€€¬l(€€€€€€€€€€€€€€€€ˆ´µ‘Àµ½É‘•ÉÌµÍØˆ°ÍÑÈ¡‘Á}½É‘•ÉÌ¤°(€€€€€€€€€€€€€€€€ˆ´µÍ¡¥Áµ•¹ÐµÑ•µÁ±…Ñ”µá±Íàˆ°ÍÑÈ¡½¹™¥}Á…Ñ  ‰‘Á}Ñ•µÁ±…Ñ”ˆ¤¤°(€€€€€€€€€€€€€€€€ˆ´µ‘…Ñ”ˆ°‘…Ñ”°(€€€€€€€€€€€€€€€€ˆ´µ½ÕÑÁÕÐµÉ½½Ðˆ°ÍÑÈ¡½ÕÑÁÕÐ¤°(€€€€€€€€€€€t°(€€€€€€€€¤(€€€•±¥˜Á±…Ñ™½É´€ôô€‰åÕ¹•áÁÉ•ÍÌˆè(€€€€€€€‘Á}½É‘•ÉÌ€ô‰…Ñ €¼€‰‘Á}½É‘•ÉÌ¹ÍØˆ(€€€€€€€åÕ¹}½É‘•ÉÌ€ô‰…Ñ €¼€‰åÕ¹}½É‘•ÉÌ¹á±Íàˆ(€€€€€€€¥˜¹½Ð‘Á}½É‘•ÉÌ¹•á¥ÍÑÌ ¤½È¹½ÐåÕ¹}½É‘•ÉÌ¹•á¥ÍÑÌ ¤è(€€€€€€€€€€€É…¥Í”Y…±Õ•ÉÉ½È ‹’êG¦S¦v‹–6W–"š.¦r¢š–#–r£–B3’â¢Ò›–>ß–º3š"Cž²°€ÇŽÈƒš¶”ˆ¤(€€€€€€€½‘”°½ÕÐ°•ÉÈ€ôÍÉ¥ÁÐ (€€€€€€€€€€€Í½ÉÑ}åÕ¹•áÁÉ•ÍÍ}±…‰•±Í}‰å}Í­Ô°(€€€€€€€€€€€l(€€€€€€€€€€€€€€€€ˆ´µé¥Àˆ°ÍÑÈ¡Í½ÕÉ”¤°(€€€€€€€€€€€€€€€€ˆ´µåÕ¹•áÁÉ•ÍÌµá±Íàˆ°ÍÑÈ¡åÕ¹}½É‘•ÉÌ¤°(€€€€€€€€€€€€€€€€ˆ´µ‘Àµ½É‘•ÉÌµÍØˆ°ÍÑÈ¡‘Á}½É‘•ÉÌ¤°(€€€€€€€€€€€€€€€€ˆ´µÍ­Ôµá±Íàˆ°ÍÑÈ¡½¹™¥}Á…Ñ  ‰Í­Ôˆ¤¤°(€€€€€€€€€€€€€€€€ˆ´µ‘…Ñ”ˆ°‘…Ñ”°(€€€€€€€€€€€€€€€€ˆ´µ½ÕÑÁÕÐµÉ½½Ðˆ°ÍÑÈ¡½ÕÑÁÕÐ¤°(€€€€€€€€€€€t°(€€€€€€€€¤(€€€•±Í”è(€€€€€€€Í™}½É‘•É}™¥±•Ì€ô±¥ÍÐ¡‰…Ñ ¹±½ˆ ‰Í™}½É‘•ÉÌ¹á±Ì¨ˆ¤¤(€€€€€€€¥˜±•¸¡Í™}½É‘•É}™¥±•Ì¤€„ô€Äè(€€€€€€€€€€€É…¥Í”Y…±Õ•ÉÉ½È ‹¦†ë’âÃ¦v‹–6Wš.–"¦r¢š–#–r£–B3’â¢Ò›–>ß–º3š"Cž²°€Èƒš¶”ˆ¤(€€€€€€€½‘”°½ÕÐ°•ÉÈ€ôÍÉ¥ÁÐ (€€€€€€€€€€€ÍÁ±¥Ñ}Í™}±…‰•±Í}‰å}Í•¹‘•È°(€€€€€€€€€€€l(€€€€€€€€€€€€€€€€ˆ´µÁ‘˜ˆ°ÍÑÈ¡Í½ÕÉ”¤°(€€€€€€€€€€€€€€€€ˆ´µÍ˜µ¥¹Ñ•É¹…Ñ¥½¹…°µ™¥±”ˆ°ÍÑÈ¡Í™}½É‘•É}™¥±•ÍlÁt¤°(€€€€€€€€€€€€€€€€ˆ´µ‘…Ñ”ˆ°‘…Ñ”°(€€€€€€€€€€€€€€€€ˆ´µ½ÕÑÁÕÐµÉ½½Ðˆ°ÍÑÈ¡½ÕÑÁÕÐ¤°(€€€€€€€€€€€t°(€€€€€€€€¤((€€€¥˜½‘”€„ô€Àè(€€€€€€€É…¥Í”Y…±Õ•ÉÉ½È¡•ÉÈ¹ÍÑÉ¥À ¤½È½ÕÐ¹ÍÑÉ¥À ¤½È€‹–’žB–’Ç¢Ò”ˆ¤(€€€¥˜½µµ¥Ñ}Ñ…É•Ðè(€€€€€€€¥˜Á±…Ñ™½É´€ôô€‰Í˜ˆ…¹ÍÑ•À€ôô€Èè(€€€€€€€€€€€™½È½±¥¸‰…Ñ ¹±½ˆ ‰Í™}½É‘•ÉÌ¹á±Ì¨ˆ¤è(€€€€€€€€€€€€€€€¥˜½±€„ô½µµ¥Ñ}Ñ…É•Ðè(€€€€€€€€€€€€€€€€€€€½±¹Õ¹±¥¹¬¡µ¥ÍÍ¥¹}½¬õQÉÕ”¤(€€€€€€€Í½ÕÉ”¹É•Á±…”¡½µµ¥Ñ}Ñ…É•Ð¤(€€€É•ÍÕ±Ð€ôÉ•ÍÕ±Ñ}™É½´¡½ÕÐ¤(€€€É•Á½ÉÐ€ôÉ•ÍÕ±Ð¹Ý¥Ñ¡}ÍÕ™™¥à ˆ¹É•Á½ÉÐ¹©Í½¸ˆ¤¥˜É•ÍÕ±Ð¹¥Í}™¥±” ¤•±Í”É•ÍÕ±Ð€¼€‹š‚‡¦ª3š"C–*š*—–F(¹©Í½¸ˆ(€€€Á…­…”€ô€ (€€€€€€€É•ÍÕ±Ð(€€€€€€€¥˜É•ÍÕ±Ð¹¥Í}™¥±” ¤(€€€€€€€•±Í”A…Ñ  (€€€€€€€€€€€Í¡ÕÑ¥°¹µ…­•}…É¡¥Ù” (€€€€€€€€€€€€€€€ÍÑÈ¡É•ÍÕ±Ð¤°(€€€€€€€€€€€€€€€€‰é¥Àˆ°(€€€€€€€€€€€€€€€É½½Ñ}‘¥ÈõÉ•ÍÕ±Ð¹Á…É•¹Ð°(€€€€€€€€€€€€€€€‰…Í•}‘¥ÈõÉ•ÍÕ±Ð¹¹…µ”°(€€€€€€€€€€€€¤(€€€€€€€€¤(€€€€¤(€€€¥˜ÍÑ•À€ôô€Ìè(€€€€€€€Í¡ÕÑ¥°¹ÉµÑÉ•”¡‰…Ñ °¥¹½É•}•ÉÉ½ÉÌõQÉÕ”¤(€€€É•ÑÕÉ¸Á…­…”°É•Á½ÉÐ¥˜É•Á½ÉÐ¹•á¥ÍÑÌ ¤•±Í”9½¹”(()…Íå¹Œ‘•˜ÁÉ½•ÍÌ (€€€É•ÅÕ•ÍÐèI•ÅÕ•ÍÐ°(€€€ÍÑ•Àè¥¹Ð°(€€€‘…Ñ”èÍÑÈ°(€€€ÕÁ±½…èUÁ±½…‘¥±”°(€€€Á±…Ñ™½É´èÍÑÈ°(€€€‰ÕÍ¥¹•ÍÍ}ÑåÁ”èÍÑÈ€ô€ˆˆ°(€€€‰…ÑÑ•ÉäèÍÑÈ€ô€‹–B˜ˆ°(¤è(€€€ÕÍ•È€ôÕÍ•É}™½È¡É•ÅÕ•ÍÐ¤(€€€¥˜¹½ÐÕÍ•Èè(€€€€€€€É•ÑÕÉ¸I•‘¥É•ÑI•ÍÁ½¹Í” ˆ½±½¥¸ˆ°ÍÑ…ÑÕÍ}½‘”ôÌÀÌ¤(€€€Í½ÕÉ”èA…Ñ ð9½¹”€ô9½¹”(€€€ÑÉäè(€€€€€€€Á±…Ñ™½É´€ô¹½Éµ…±¥é•}Á±…Ñ™½É´¡Á±…Ñ™½É´¤(€€€€€€€‘…Ñ”€ô¹½Éµ…±¥é•}‘…Ñ”¡‘…Ñ”¤(€€€€€€€É•ÅÕ¥É•€ôÉ•ÅÕ¥É•‘}½¹™¥}­•åÌ¡ÍÑ•À°Á±…Ñ™½É´¤(€€€€€€€µ¥ÍÍ¥¹œ€ôm=9%}1	1Mm­•åt™½È­•ä¥¸É•ÅÕ¥É•¥˜¹½Ð½¹™¥}Á…Ñ ¡­•ä¤¹•á¥ÍÑÌ ¥t(€€€€€€€¥˜µ¥ÍÍ¥¹œè(€€€€€€€€€€€É…¥Í”Y…±Õ•ÉÉ½È ‹¢¾ß–#–r£–në–ºk¦7žö»’â·’â+’òƒ¾òhˆ€¬€‹Žˆ¹©½¥¸¡µ¥ÍÍ¥¹œ¤¤((€€€€€€€‰…Ñ €ô‰…Ñ¡}‘¥È¡ÕÍ•Él‰¥‰t°Á±…Ñ™½É´¤(€€€€€€€•áÑ•¹Í¥½¹Ì€ôì(€€€€€€€€€€€€ Ä°€‰åÕ¹•áÁÉ•ÍÌˆ¤èìˆ¹ÍØ‰ô°(€€€€€€€€€€€€ Ä°€‰Í˜ˆ¤èìˆ¹ÍØ‰ô°(€€€€€€€€€€€€ È°€‰åÕ¹•áÁÉ•ÍÌˆ¤èìˆ¹á±Íà‰ô°(€€€€€€€€€€€€ È°€‰Í˜ˆ¤èìˆ¹á±Ìˆ°€ˆ¹á±Íà‰ô°(€€€€€€€€€€€€ Ì°€‰åÕ¹•áÁÉ•ÍÌˆ¤èìˆ¹é¥À‰ô°(€€€€€€€€€€€€ Ì°€‰Í˜ˆ¤èìˆ¹Á‘˜‰ô°(€€€€€€€ô(€€€€€€€Í½ÕÉ”€ô…Ý…¥ÐÍ…Ù•}ÕÁ±½…¡ÕÁ±½…°‰…Ñ °•áÑ•¹Í¥½¹Íl¡ÍÑ•À°Á±…Ñ™½É´¥t¤(€€€€€€€…Íå¹ŒÝ¥Ñ ]=I-1=]}1=,è(€€€€€€€€€€€Á…­…”°É•Á½ÉÐ€ô…Ý…¥ÐÉÕ¹}¥¹}Ñ¡É•…‘Á½½° (€€€€€€€€€€€€€€€•á•ÕÑ•}Ý½É­™±½Ü°(€€€€€€€€€€€€€€€ÕÍ•Él‰¥‰t°(€€€€€€€€€€€€€€€ÍÑ•À°(€€€€€€€€€€€€€€€‘…Ñ”°(€€€€€€€€€€€€€€€Í½ÕÉ”°(€€€€€€€€€€€€€€€Á±…Ñ™½É´°(€€€€€€€€€€€€€€€‰ÕÍ¥¹•ÍÍ}ÑåÁ”°(€€€€€€€€€€€€€€€‰…ÑÑ•Éä°(€€€€€€€€€€€€¤(€€€€€€€ÍÑ•Á}¹…µ”€ô˜‹š¶—¦ªíÍÑ•Áôƒ
ÜíA1Q=I5}1	1MmÁ±…Ñ™½Éµuôˆ(€€€€€€€…‘‘}ÉÕ¸¡ÕÍ•Él‰¥‰t°ÍÑ•Á}¹…µ”°‘…Ñ”°€‰MUMLˆ°Á…­…”°É•Á½ÉÐ¤(€€€€€€€É•ÑÕÉ¸¥±•I•ÍÁ½¹Í”¡Á…­…”°™¥±•¹…µ”õÁ…­…”¹¹…µ”¤(€€€•á•ÁÐá•ÁÑ¥½¸…Ì•áŒè(€€€€€€€¥˜Í½ÕÉ”…¹Í½ÕÉ”¹•á¥ÍÑÌ ¤è(€€€€€€€€€€€Í½ÕÉ”¹Õ¹±¥¹¬¡µ¥ÍÍ¥¹}½¬õQÉÕ”¤(€€€€€€€±…‰•°€ôA1Q=I5}1	1L¹•Ð¡Á±…Ñ™½É´°€‹šr«ž~—–æÏ–>Àˆ¤¥˜¥Í¥¹ÍÑ…¹”¡Á±…Ñ™½É´°ÍÑÈ¤•±Í”€‹šr«ž~—–æÏ–>Àˆ(€€€€€€€…‘‘}ÉÕ¸¡ÕÍ•Él‰¥‰t°˜‹š¶—¦ªíÍÑ•Áôƒ
Üí±…‰•±ôˆ°‘…Ñ”°€‰%1ˆ°9½¹”°9½¹”°ÍÑÈ¡•áŒ¤¤(€€€€€€€É•ÑÕÉ¸!Q51I•ÍÁ½¹Í”¡Á…” ‹–’žB–’Ç¢Ò”ˆ°˜œñÍ•Ñ¥½¸±…ÍÌô‰±½¥¸ˆøñ Äû–’žB–’Ç¢Ò”ð½ ÄøñÀ±…ÍÌô‰±•…ˆùí¡Ñµ°¹•Í…Á”¡ÍÑÈ¡•áŒ¤¥ôð½Àøñ„±…ÍÌô‰‰ÕÑÑ½¸ˆ¡É•˜ôˆ¼ˆû¢þS–n{–Þ—’ös–>Àð½„øð½Í•Ñ¥½¸øœ°ÕÍ•È¤°ÍÑ…ÑÕÍ}½‘”ôÐÀÀ¤(()…ÁÀ¹Á½ÍÐ ˆ½ÍÑ•À¼Äˆ¤)…Íå¹Œ‘•˜ÍÑ•ÀÄ (€€€É•ÅÕ•ÍÐèI•ÅÕ•ÍÐ°(€€€‘…Ñ”èÍÑÈ€ô½É´ ¸¸¸¤°(€€€Á±…Ñ™½É´èÍÑÈ€ô½É´ ¸¸¸¤°(€€€‰ÕÍ¥¹•ÍÍ}ÑåÁ”èÍÑÈ€ô½É´ ˆˆ¤°(€€€‰…ÑÑ•ÉäèÍÑÈ€ô½É´ ‹–B˜ˆ¤°(€€€™¥±”èUÁ±½…‘¥±”€ô¥±” ¸¸¸¤°(¤è(€€€É•ÑÕÉ¸…Ý…¥ÐÁÉ½•ÍÌ¡É•ÅÕ•ÍÐ°€Ä°‘…Ñ”°™¥±”°Á±…Ñ™½É´°‰ÕÍ¥¹•ÍÍ}ÑåÁ”°‰…ÑÑ•Éä¤()…ÁÀ¹Á½ÍÐ ˆ½ÍÑ•À¼Èˆ¤)…Íå¹Œ‘•˜ÍÑ•ÀÈ¡É•ÅÕ•ÍÐèI•ÅÕ•ÍÐ°‘…Ñ”èÍÑÈ€ô½É´ ¸¸¸¤°Á±…Ñ™½É´èÍÑÈ€ô½É´ ¸¸¸¤°™¥±”èUÁ±½…‘¥±”€ô¥±” ¸¸¸¤¤è(€€€É•ÑÕÉ¸…Ý…¥ÐÁÉ½•ÍÌ¡É•ÅÕ•ÍÐ°€È°‘…Ñ”°™¥±”°Á±…Ñ™½É´¤()…ÁÀ¹Á½ÍÐ ˆ½ÍÑ•À¼Ìˆ¤)…Íå¹Œ‘•˜ÍÑ•ÀÌ¡É•ÅÕ•ÍÐèI•ÅÕ•ÍÐ°‘…Ñ”èÍÑÈ€ô½É´ ¸¸¸¤°Á±…Ñ™½É´èÍÑÈ€ô½É´ ¸¸¸¤°™¥±”èUÁ±½…‘¥±”€ô¥±” ¸¸¸¤¤è(€€€É•ÑÕÉ¸…Ý…¥ÐÁÉ½•ÍÌ¡É•ÅÕ•ÍÐ°€Ì°‘…Ñ”°™¥±”°Á±…Ñ™½É´¤(()…ÁÀ¹•Ð ˆ½¡•…±Ñ¡èˆ¤)‘•˜¡•…±Ñ¡è ¤è(€€€É•ÑÕÉ¸ì‰ÍÑ…ÑÕÌˆè€‰½¬ˆ°€‰Á±…Ñ™½ÉµÌˆè±¥ÍÐ¡A1Q=I5}1	1L¥ô()…ÁÀ¹•Ð ˆ½‘½Ý¹±½…½íÉÕ¹}¥‘ôˆ¤)‘•˜‘½Ý¹±½…¡É•ÅÕ•ÍÐèI•ÅÕ•ÍÐ°ÉÕ¹}¥è¥¹Ð¤è(€€€ÕÍ•È€ôÕÍ•É}™½È¡É•ÅÕ•ÍÐ¤(€€€¥˜¹½ÐÕÍ•ÈèÉ•ÑÕÉ¸I•‘¥É•ÑI•ÍÁ½¹Í” ˆ½±½¥¸ˆ°ÍÑ…ÑÕÍ}½‘”ôÌÀÌ¤(€€€Ý¥Ñ ±½Í¥¹œ¡½¹¹•Ñ¥½¸ ¤¤…Ì½¹¸è(€€€€€€€¥˜ÕÍ•Él‰É½±”‰t€ôô€‰…‘µ¥¸ˆè(€€€€€€€€€€€É½Ü€ô½¹¸¹•á•ÕÑ” ‰M1PÉ•ÍÕ±Ñ}Á…Ñ I=4ÉÕ¹Ì]!I¥ôüˆ°€¡ÉÕ¹}¥°¤¤¹™•Ñ¡½¹” ¤(€€€€€€€•±Í”è(€€€€€€€€€€€É½Ü€ô½¹¸¹•á•ÕÑ” (€€€€€€€€€€€€€€€€‰M1PÉ•ÍÕ±Ñ}Á…Ñ I=4ÉÕ¹Ì]!I¥ôü9ÕÍ•É}¥ôüˆ°(€€€€€€€€€€€€€€€€¡ÉÕ¹}¥°ÕÍ•Él‰¥‰t¤°(€€€€€€€€€€€€¤¹™•Ñ¡½¹” ¤(€€€¥˜¹½ÐÉ½Ü½È¹½ÐÉ½Ýl‰É•ÍÕ±Ñ}Á…Ñ ‰t½È¹½ÐA…Ñ ¡É½Ýl‰É•ÍÕ±Ñ}Á…Ñ ‰t¤¹•á¥ÍÑÌ ¤èÉ•ÑÕÉ¸!Q51I•ÍÁ½¹Í” ‹žîOšzs’â7–¶c–r ˆ°ÍÑ…ÑÕÍ}½‘”ôÐÀÐ¤(€€€Á…Ñ €ôA…Ñ ¡É½Ýl‰É•ÍÕ±Ñ}Á…Ñ ‰t¤ìÉ•ÑÕÉ¸¥±•I•ÍÁ½¹Í”¡Á…Ñ °™¥±•¹…µ”õÁ…Ñ ¹¹…µ”¤(()ML€ôÈœœœ(éÉ½½Ñí™½¹Ðµ™…µ¥±äèµ…ÁÁ±”µÍåÍÑ•´±	±¥¹­5…MåÍÑ•µ½¹Ð°‰MAÉ¼¥ÍÁ±…äˆ°‰A¥¹…¹œMˆ°‰5¥É½Í½™Ðe…!•¤U$ˆ±Í…¹ÌµÍ•É¥˜í½±½ÈèŒÄàÈÀÉ„í‰…­É½Õ¹è••˜Å˜Ôí±•ÑÑ•ÈµÍÁ…¥¹œèÁô©í‰½àµÍ¥é¥¹œé‰½É‘•Èµ‰½àí±•ÑÑ•ÈµÍÁ…¥¹œèÁõ‰½‘åíµ…É¥¸èÀíµ¥¸µ¡•¥¡ÐèÄÀÁÙ í‰…­É½Õ¹è••˜Å˜Õõ…íÑ•áÐµ‘•½É…Ñ¥½¸é¹½¹”í½±½Èé¥¹¡•É¥Ñõ¡•…‘•Éí¡•¥¡ÐèØÙÁàíÁ…‘‘¥¹œèÀµ…à ÈÑÁà°ÑÙÜ¤í‘¥ÍÁ±…äé™±•àí…±¥¸µ¥Ñ•µÌé•¹Ñ•Èí©ÕÍÑ¥™äµ½¹Ñ•¹ÐéÍÁ…”µ‰•ÑÝ••¸í‰…­É½Õ¹éÉ‰„ ÈÔÀ°ÈÔÄ°ÈÔÌ°¸àØ¤í‰½É‘•Èµ‰½ÑÑ½´èÅÁàÍ½±¥É‰„ ØÐ°ÜÌ°àà°¸ÄØ¤í‰…­‘É½Àµ™¥±Ñ•Èé‰±ÕÈ ÈÁÁà¤íÁ½Í¥Ñ¥½¸éÍÑ¥­äíÑ½ÀèÀíèµ¥¹‘•àèÍô¹‰É…¹‘í™½¹ÐµÍ¥é”èÄáÁàí™½¹ÐµÝ•¥¡ÐèÜØÁô¹‰É…¹Íµ…±±í™½¹ÐµÍ¥é”èåÁàí½±½ÈèŒÜààÈäÄíµ…É¥¸µ±•™ÐèáÁáõ¹…Ùí‘¥ÍÁ±…äé™±•àí…ÀèÈÉÁàí½±½ÈèŒØØÜÄàÄí™½¹ÐµÍ¥é”èÄÍÁàí™½¹ÐµÝ•¥¡ÐèØÔÁõ¹…Ø„é¡½Ù•Éí½±½ÈèŒÀÀÜÅ”Íõµ…¥¹íÝ¥‘Ñ éµ¥¸ ÄÈÀÁÁà±…±Œ ÄÀÀ”€´€ÌÉÁà¤¤íµ…É¥¸é…ÕÑ¼íÁ…‘‘¥¹œèÌÑÁà€À€ØÑÁáô¹¡•É½í‘¥ÍÁ±…äé™±•àí©ÕÍÑ¥™äµ½¹Ñ•¹ÐéÍÁ…”µ‰•ÑÝ••¸í…±¥¸µ¥Ñ•µÌé•¹í…ÀèÈáÁàíµ…É¥¸µ‰½ÑÑ½´èÈÑÁáô¹•å•‰É½Ýí™½¹ÐµÍ¥é”èÄÁÁàí½±½ÈèŒÜÈÝáí™½¹ÐµÝ•¥¡ÐèÜÔÀíµ…É¥¸èÀ€À€åÁáô¹¡•É¼ Ä°¹¡•… Åí™½¹ÐµÍ¥é”èÌÑÁàí±¥¹”µ¡•¥¡ÐèÄ¸Ààíµ…É¥¸èÀ€À€ÄÅÁáô¹±•…‘í½±½ÈèŒØÈÙÝŒí±¥¹”µ¡•¥¡ÐèÄ¸ØÔíµ…àµÝ¥‘Ñ èÜÄÁÁàíµ…É¥¸èÀí™½¹ÐµÍ¥é”èÄÑÁáô¹¥‘•¹Ñ¥Ñåí‘¥ÍÁ±…äé™±•àí…ÀèÄÁÁàí…±¥¸µ¥Ñ•µÌé•¹Ñ•Èí‰…­É½Õ¹éÉ‰„ ÈÔÔ°ÈÔÔ°ÈÔÔ°¸ÜÈ¤íÁ…‘‘¥¹œèÄÁÁà€ÄÍÁàí‰½É‘•ÈèÅÁàÍ½±¥É‰„ ØÜ°ÜÜ°äÈ°¸ÄÔ¤í‰½É‘•ÈµÉ…‘¥ÕÌèáÁáô¹¥‘•¹Ñ¥Ñä¥íÝ¥‘Ñ èáÁàí¡•¥¡ÐèáÁàí‰½É‘•ÈµÉ…‘¥ÕÌèÔÀ”í‰…­É½Õ¹èŒÈÍ„ÔØØí‰½àµÍ¡…‘½ÜèÀ€À€À€ÍÁà€á˜É”Ñô¹¥‘•¹Ñ¥Ñäˆ°¹¥‘•¹Ñ¥ÑäÍµ…±±í‘¥ÍÁ±…äé‰±½­ô¹¥‘•¹Ñ¥Ñä‰í™½¹ÐµÍ¥é”èÄÉÁáô¹¥‘•¹Ñ¥ÑäÍµ…±±í™½¹ÐµÍ¥é”èÄÅÁàí½±½ÈèŒÜÜàÈäÌíµ…É¥¸µÑ½ÀèÍÁáô¹½¹™¥í‘¥ÍÁ±…äé™±•àí…±¥¸µ¥Ñ•µÌé•¹Ñ•Èí©ÕÍÑ¥™äµ½¹Ñ•¹ÐéÍÁ…”µ‰•ÑÝ••¸íÁ…‘‘¥¹œèÄÕÁà€ÄÝÁàí‰…­É½Õ¹éÉ‰„ ÈÔÔ°ÈÔÔ°ÈÔÔ°¸ÜÐ¤í‰½É‘•ÈèÅÁàÍ½±¥É‰„ ØÐ°ÜÐ°àä°¸ÄÐ¤í‰½É‘•ÈµÉ…‘¥ÕÌèáÁàíµ…É¥¸µ‰½ÑÑ½´èÄÑÁáô¹½¹™¥œÍµ…±±í½±½ÈèŒÜÌÝ”áíµ…É¥¸µÉ¥¡ÐèÄÉÁáô¹½¹™¥œ‰í™½¹ÐµÍ¥é”èÄÑÁáô¹½¹™¥œ•µí‘¥ÍÁ±…äé‰±½¬í½±½ÈèŒÝ„àÔäÔí™½¹ÐµÍ¥é”èÄÅÁàí™½¹ÐµÍÑå±”é¹½Éµ…°íµ…É¥¸µÑ½ÀèÑÁáô¹ÍÑ•ÁÍí‘¥ÍÁ±…äéÉ¥íÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹ÌéÉ•Á•…Ð Ì±µ¥¹µ…à À°Å™È¤¤í…ÀèÄÉÁáô¹ÍÑ•ÁÌ…ÉÑ¥±”°¹Á…¹•°°¹¡¥ÍÑ½Éåí‰…­É½Õ¹éÉ‰„ ÈÔÔ°ÈÔÔ°ÈÔÔ°¸Üà¤í‰½É‘•ÈèÅÁàÍ½±¥É‰„ ØÐ°ÜÐ°àä°¸ÄÐ¤í‰½É‘•ÈµÉ…‘¥ÕÌèáÁàíÁ…‘‘¥¹œèÄáÁàí‰½àµÍ¡…‘½ÜèÀ€ÄÁÁà€ÈáÁàÉ‰„ ÐÄ°ÔÀ°ØØ°¸ÀØ¥ô¹ÍÑ•Àµ¡•…‘í¡•¥¡ÐèÈÕÁàí‘¥ÍÁ±…äé™±•àí…±¥¸µ¥Ñ•µÌé•¹Ñ•Èí©ÕÍÑ¥™äµ½¹Ñ•¹ÐéÍÁ…”µ‰•ÑÝ••¸íµ…É¥¸µ‰½ÑÑ½´èÄÍÁáô¹ÍÑ•Àµ¡•…ùÍÁ…¹í™½¹ÐµÍ¥é”èÄÁÁàí½±½ÈèŒÝ„àÐäÈí™½¹ÐµÝ•¥¡ÐèÜÀÁô¹¹½íÝ¥‘Ñ èÈÕÁàí¡•¥¡ÐèÈÕÁàí‘¥ÍÁ±…äéÉ¥íÁ±…”µ¥Ñ•µÌé•¹Ñ•Èí‰½É‘•ÈµÉ…‘¥ÕÌèÝÁàí‰…­É½Õ¹è”Õ˜Å™˜í½±½ÈèŒÀÀÙ•‘ˆí™½¹ÐµÝ•¥¡ÐèàÀÀí™½¹ÐµÍ¥é”èÄÅÁáô¹ÍÑ•ÁÌ È°¹Á…¹•° È°¹¡¥ÍÑ½Éä Éí™½¹ÐµÍ¥é”èÄÝÁàíµ…É¥¸èÀ€À€ÝÁáô¹ÍÑ•ÁÌù…ÉÑ¥±”ùÁí¡•¥¡ÐèÐÉÁàí½±½ÈèŒÙ”ÜààÜí™½¹ÐµÍ¥é”èÄÉÁàí±¥¹”µ¡•¥¡ÐèÄ¸ÔÔíµ…É¥¸èÀ€À€ÄÍÁáô¹Ý½É­™±½Üµ™½Éµí‘¥ÍÁ±…äéÉ¥í…ÀèÄÁÁáô¹Á±…Ñ™½É´µÁ¥­•È°¹µ¥¹¤µÍ•µ•¹Ñí‘¥ÍÁ±…äéÉ¥íÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹ÌèÅ™È€Å™Èí‰…­É½Õ¹è”å•‘˜Èí‰½É‘•ÈèÅÁàÍ½±¥€‘”Å”Üí‰½É‘•ÈµÉ…‘¥ÕÌèáÁàíÁ…‘‘¥¹œèÍÁàí…ÀèÍÁáô¹Á±…Ñ™½É´µÁ¥­•È±…‰•°°¹µ¥¹¤µÍ•µ•¹Ð±…‰•±íÕÉÍ½ÈéÁ½¥¹Ñ•Éô¹Á±…Ñ™½É´µÁ¥­•È¥¹ÁÕÐ°¹µ¥¹¤µÍ•µ•¹Ð¥¹ÁÕÑíÁ½Í¥Ñ¥½¸é…‰Í½±ÕÑ”í½Á…¥ÑäèÀíÁ½¥¹Ñ•Èµ•Ù•¹ÑÌé¹½¹•ô¹Á±…Ñ™½É´µÁ¥­•ÈÍÁ…¸°¹µ¥¹¤µÍ•µ•¹ÐÍÁ…¹í¡•¥¡ÐèÌÁÁàí‘¥ÍÁ±…äéÉ¥íÁ±…”µ¥Ñ•µÌé•¹Ñ•Èí‰½É‘•ÈµÉ…‘¥ÕÌèÙÁàí½±½ÈèŒØØÜÄàÄí™½¹ÐµÍ¥é”èÄÅÁàí™½¹ÐµÝ•¥¡ÐèÜÀÁô¹Á±…Ñ™½É´µÁ¥­•È¥¹ÁÕÐé¡•­•­ÍÁ…¸°¹µ¥¹¤µÍ•µ•¹Ð¥¹ÁÕÐé¡•­•­ÍÁ…¹í‰…­É½Õ¹è™™˜í½±½ÈèŒÀàÙ‘Œí‰½àµÍ¡…‘½ÜèÀ€ÅÁà€ÑÁàÉ‰„ ÌÜ°ÐÔ°Ôä°¸ÄÐ¥õ…ÉÑ¥±•m‘…Ñ„µÁ±…Ñ™½É´ô‰åÕ¹•áÁÉ•ÍÌ‰t€¹Í˜µ½¹±åí‘¥ÍÁ±…äé¹½¹•õ…ÉÑ¥±•m‘…Ñ„µÁ±…Ñ™½É´ô‰Í˜‰t€¹åÕ¸µ½¹±åí‘¥ÍÁ±…äé¹½¹•ô¹™¥±•í‘¥ÍÁ±…äé‰±½¬í‰½É‘•ÈèÅÁà‘…Í¡•€ˆáŒÉ˜í‰½É‘•ÈµÉ…‘¥ÕÌèáÁàíÁ…‘‘¥¹œèÄÅÁà€ÄÉÁàí‰…­É½Õ¹è˜á™…™ŒíÕÉÍ½ÈéÁ½¥¹Ñ•Èí½Ù•É™±½Üé¡¥‘‘•¹ô¹™¥±”é¡½Ù•Éí‰½É‘•Èµ½±½ÈèŒÝ„åÜí‰…­É½Õ¹è˜Ñ˜á™ô¹™¥±”‰í‘¥ÍÁ±…äé‰±½¬í™½¹ÐµÍ¥é”èÄÉÁàíµ…É¥¸µ‰½ÑÑ½´èÝÁáô¹™¥±”ùÍÁ…¸é±…ÍÐµ¡¥±‘í‘¥ÍÁ±…äé‰±½¬í½±½ÈèŒÜÌàÀäÈí™½¹ÐµÍ¥é”èÄÅÁàíÝ¡¥Ñ”µÍÁ…”é¹½ÝÉ…Àí½Ù•É™±½Üé¡¥‘‘•¸íÑ•áÐµ½Ù•É™±½Üé•±±¥ÁÍ¥Íô¹™¥±”¥¹ÁÕÑí‘¥ÍÁ±…äé¹½¹•ô¹Í˜µ½ÁÑ¥½¹Íí‰½É‘•ÈèÅÁàÍ½±¥€‘”É”äí‰…­É½Õ¹è˜Ý˜å™ˆí‰½É‘•ÈµÉ…‘¥ÕÌèáÁàíÁ…‘‘¥¹œèÄÁÁàí…ÀèåÁáô¹Í˜µ½ÁÑ¥½¹Ìù±…‰•±í‘¥ÍÁ±…äéÉ¥í…ÀèÙÁàí½±½ÈèŒØàÜÐàØí™½¹ÐµÍ¥é”èÄÅÁàí™½¹ÐµÝ•¥¡ÐèÜÀÁô¹Í˜µ½ÁÑ¥½¹ÌÍ•±•ÑíÝ¥‘Ñ èÄÀÀ•ô¹Í˜µ½ÁÑ¥½¹Ìù‘¥Ùí‘¥ÍÁ±…äéÉ¥íÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹Ìé…ÕÑ¼€ÄÈÁÁàí…±¥¸µ¥Ñ•µÌé•¹Ñ•Èí…ÀèÄÁÁáô¹Í˜µ½ÁÑ¥½¹Ìù‘¥Øù‰í™½¹ÐµÍ¥é”èÄÅÁàí½±½ÈèŒØàÜÐàÙô¹µ¥¹¤µÍ•µ•¹ÐÍÁ…¹í¡•¥¡ÐèÈÝÁáô¹É½Ýí‘¥ÍÁ±…äéÉ¥íÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹Ìéµ¥¹µ…à À°Å™È¤…ÕÑ¼í…ÀèáÁáô¹É½Ü¥¹ÁÕÐ±¥¹ÁÕÐ±Í•±•Ñíµ¥¸µÝ¥‘Ñ èÀí‰½É‘•ÈèÅÁàÍ½±¥€‘Ñ‘í‰…­É½Õ¹è™™˜í‰½É‘•ÈµÉ…‘¥ÕÌèÝÁàíÁ…‘‘¥¹œèåÁà€ÄÁÁàí½±½ÈèŒÈÀÉ„ÌØí™½¹Ðé¥¹¡•É¥Ðí™½¹ÐµÍ¥é”èÄÉÁáô¹É½Ü¥¹ÁÕÐé™½ÕÌ±¥¹ÁÕÐé™½ÕÌ±Í•±•Ðé™½ÕÍí½ÕÑ±¥¹”èÉÁàÍ½±¥É‰„ À°ÄÄÌ°ÈÈÜ°¸È¤í‰½É‘•Èµ½±½ÈèŒÐÜäÕ‘ô¹ÁÉ¥µ…Éä°¹‰ÕÑÑ½¹í‰½É‘•ÈèÀí‰½É‘•ÈµÉ…‘¥ÕÌèÝÁàíÁ…‘‘¥¹œèåÁà€ÄÍÁàí™½¹Ðé¥¹¡•É¥Ðí™½¹ÐµÝ•¥¡ÐèÜÈÀí™½¹ÐµÍ¥é”èÄÉÁàíÕÉÍ½ÈéÁ½¥¹Ñ•Èí‘¥ÍÁ±…äé¥¹±¥¹”µÉ¥íÁ±…”µ¥Ñ•µÌé•¹Ñ•Éô¹ÁÉ¥µ…Éåí‰…­É½Õ¹èŒÀàÜá‘˜í½±½Èè™™™ô¹ÁÉ¥µ…Éäé¡½Ù•Éí‰…­É½Õ¹èŒÀÀÙ‘ô¹Í½™Ñí‰…­É½Õ¹è”á˜É™Œí½±½ÈèŒÄÈØÝˆÕô¹¡¥ÍÑ½Éåíµ…É¥¸µÑ½ÀèÄÑÁáô¹Í•Ñ¥½¸µ¡•…‘í‘¥ÍÁ±…äé™±•àí©ÕÍÑ¥™äµ½¹Ñ•¹ÐéÍÁ…”µ‰•ÑÝ••¸í…±¥¸µ¥Ñ•µÌé•¹íµ…É¥¸µ‰½ÑÑ½´èÄÁÁáô¹Í•Ñ¥½¸µ¡•…ùÍÁ…¹í½±½ÈèŒÝàÜäÐí™½¹ÐµÍ¥é”èÄÅÁáô¹Ñ…‰±”µÝÉ…Áí½Ù•É™±½Üµàé…ÕÑ½õÑ…‰±•íÝ¥‘Ñ èÄÀÀ”í‰½É‘•Èµ½±±…ÁÍ”é½±±…ÁÍ”í™½¹ÐµÍ¥é”èÄÉÁáõÑ¡íÑ•áÐµ…±¥¸é±•™Ðí½±½ÈèŒÝàÜäÐí™½¹ÐµÍ¥é”èÄÁÁáõÑ ±Ñ‘íÁ…‘‘¥¹œèÄÅÁà€áÁàí‰½É‘•Èµ‰½ÑÑ½´èÅÁàÍ½±¥€”Í”Ý•õÑ‘í½±½ÈèŒÐÔÔÄØÑô¹ÍÑ…ÑÕÍí™½¹ÐµÍ¥é”èÄÁÁàí™½¹ÐµÝ•¥¡ÐèàÀÀíÁ…‘‘¥¹œèÑÁà€ÝÁàí‰½É‘•ÈµÉ…‘¥ÕÌèÙÁáô¹ÍÑ…ÑÕÌ¹ÍÕ•ÍÍí‰…­É½Õ¹è”Í˜Õ•ˆí½±½ÈèŒÈÌÝ„ÔÁô¹ÍÑ…ÑÕÌ¹™…¥±•‘í‰…­É½Õ¹è™”å”àí½±½Èè…ÐàÐÑõÑ…í½±½ÈèŒÀàÜÙ‘„í™½¹ÐµÝ•¥¡ÐèÜÀÁô¹•µÁÑåíÑ•áÐµ…±¥¸é•¹Ñ•Èí½±½ÈèŒàÜäÄå˜íÁ…‘‘¥¹œèÈÙÁáô¹¡•…‘íµ…É¥¸µ‰½ÑÑ½´èÈÉÁáô¹ÑÝ½í‘¥ÍÁ±…äéÉ¥íÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹ÌèÄ¸ÄÕ™È€¸àÕ™Èí…ÀèÄÑÁáô¹Á…¹•°™½Éµí‘¥ÍÁ±…äéÉ¥í…ÀèÄÉÁáô¹Á…¹•°±…‰•°°¹±½¥¸±…‰•±í‘¥ÍÁ±…äéÉ¥í…ÀèÝÁàí½±½ÈèŒØÌÜÀàÌí™½¹ÐµÍ¥é”èÄÉÁàí™½¹ÐµÝ•¥¡ÐèÜÀÁô¹Á…¹•°¥¹ÁÕÐ°¹±½¥¸¥¹ÁÕÑíÝ¥‘Ñ èÄÀÀ•ô¹ÅÕ¥•Ñí‰…­É½Õ¹éÉ‰„ ÈÐÜ°ÈÐä°ÈÔÈ°¸àØ¥ô¹ÅÕ¥•ÐÕ±íÁ…‘‘¥¹œèÀíµ…É¥¸èÄÁÁà€À€ÈÉÁàí±¥ÍÐµÍÑå±”é¹½¹”í‘¥ÍÁ±…äéÉ¥í…ÀèÄÑÁáô¹ÅÕ¥•Ð±¥í‘¥ÍÁ±…äé™±•àí…ÀèÄÁÁàí…±¥¸µ¥Ñ•µÌé•¹Ñ•Éô¹ÅÕ¥•Ð±¤¥íÝ¥‘Ñ èÝÁàí¡•¥¡ÐèÝÁàí‰½É‘•ÈµÉ…‘¥ÕÌèÔÀ”í‰…­É½Õ¹èŒÀàÜá‘™ô¹ÅÕ¥•Ð±¤ˆ°¹ÅÕ¥•Ð±¤Íµ…±±í‘¥ÍÁ±…äé‰±½­ô¹ÅÕ¥•Ð±¤‰í™½¹ÐµÍ¥é”èÄÍÁáô¹ÅÕ¥•Ð±¤Íµ…±°°¹¡¥¹Ñí™½¹ÐµÍ¥é”èÄÅÁàí½±½ÈèŒÜààÐäÑô¹¡¥¹Ñí±¥¹”µ¡•¥¡ÐèÄ¸ÔÔíµ…É¥¸èÁô¹¹½Ñ¥•íÁ…‘‘¥¹œèÄÑÁàí‰…­É½Õ¹è˜Å˜Ñ˜àí‰½É‘•ÈèÅÁàÍ½±¥€‘”É”äí‰½É‘•ÈµÉ…‘¥ÕÌèáÁàí½±½ÈèŒØØÜÌàÔí™½¹ÐµÍ¥é”èÄÉÁàí±¥¹”µ¡•¥¡ÐèÄ¸Ùô¹±½¥¹íµ…àµÝ¥‘Ñ èÐÄÁÁàíµ…É¥¸èÝÙ …ÕÑ¼€Àí‰…­É½Õ¹éÉ‰„ ÈÔÔ°ÈÔÔ°ÈÔÔ°¸àÈ¤íÁ…‘‘¥¹œèÌÁÁàí‰½É‘•ÈèÅÁàÍ½±¥É‰„ ØÐ°ÜÐ°àä°¸ÄÐ¤í‰½É‘•ÈµÉ…‘¥ÕÌèáÁàí‰½àµÍ¡…‘½ÜèÀ€ÄáÁà€ÐÑÁàÉ‰„ ÐÄ°ÔÀ°ØØ°¸Àä¥ô¹µ…É­í™½¹ÐµÍ¥é”èÈÍÁàí™½¹ÐµÝ•¥¡ÐèàÀÀíµ…É¥¸µ‰½ÑÑ½´èÌÁÁáô¹µ…É¬ÍÁ…¹í½±½ÈèŒÀàÜá‘™ô¹±½¥¸ Åí™½¹ÐµÍ¥é”èÈÝÁàíµ…É¥¸èÀ€À€ÄÁÁáô¹±½¥¸™½Éµí‘¥ÍÁ±…äéÉ¥í…ÀèÄÍÁàíµ…É¥¸µÑ½ÀèÈÉÁáõµ•‘¥„¡µ…àµÝ¥‘Ñ èäàÁÁà¥ì¹ÍÑ•ÁÍíÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹ÌèÅ™Éô¹ÍÑ•ÁÌù…ÉÑ¥±”ùÁí¡•¥¡Ðé…ÕÑ¼íµ¥¸µ¡•¥¡ÐèÁô¹ÑÝ½íÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹ÌèÅ™Éõõµ•‘¥„¡µ…àµÝ¥‘Ñ èØÈÁÁà¥í¡•…‘•Éí¡•¥¡Ðé…ÕÑ¼íµ¥¸µ¡•¥¡ÐèØÉÁàíÁ…‘‘¥¹œèÄÍÁà€ÄÙÁàí…±¥¸µ¥Ñ•µÌé™±•àµÍÑ…ÉÐí…ÀèÄÉÁáô¹‰É…¹Íµ…±±í‘¥ÍÁ±…äé¹½¹•õ¹…Ùí…ÀèÄÉÁàí™±•àµÝÉ…ÀéÝÉ…Àí©ÕÍÑ¥™äµ½¹Ñ•¹Ðé™±•àµ•¹‘õµ…¥¹íÝ¥‘Ñ éµ¥¸ ÄÀÀ”€´€ÈÁÁà°ÄÈÀÁÁà¤íÁ…‘‘¥¹œµÑ½ÀèÈÉÁáô¹¡•É½í…±¥¸µ¥Ñ•µÌéÍÑ…ÉÐí™±•àµ‘¥É•Ñ¥½¸é½±Õµ¹ô¹¡•É¼ Ä°¹¡•… Åí™½¹ÐµÍ¥é”èÈåÁáô¹¥‘•¹Ñ¥ÑåíÝ¥‘Ñ èÄÀÀ•ô¹½¹™¥í…±¥¸µ¥Ñ•µÌé™±•àµÍÑ…ÉÐí…ÀèÄÉÁáô¹É½ÝíÉ¥µÑ•µÁ±…Ñ”µ½±Õµ¹ÌèÅ™Éô¹É½Ü€¹ÁÉ¥µ…ÉåíÝ¥‘Ñ èÄÀÀ•ô¹Í•Ñ¥½¸µ¡•…‘í…±¥¸µ¥Ñ•µÌéÍÑ…ÉÐí…ÀèÄÁÁàí™±•àµ‘¥É•Ñ¥½¸é½±Õµ¹õô(œœœ()MI%AP€ôÈœœœ)½¹ÍÐ±½…±Q½‘…ä€ô¹•Ü…Ñ”¡…Ñ”¹¹½Ü ¤€´¹•Ü…Ñ” ¤¹•ÑQ¥µ•é½¹•=™™Í•Ð ¤€¨€ØÀÀÀÀ¤(€€¹Ñ½%M=MÑÉ¥¹œ ¤(€€¹Í±¥” À°€ÄÀ¤ì)‘½Õµ•¹Ð¹ÅÕ•ÉåM•±•Ñ½É±° ¥¹ÁÕÑmÑåÁ”ô‰‘…Ñ”‰tœ¤¹™½É…  ¡¥¹ÁÕÐ¤€ôøì(€¥¹ÁÕÐ¹Ù…±Õ”€ô±½…±Q½‘…äì)ô¤ì()‘½Õµ•¹Ð¹ÅÕ•ÉåM•±•Ñ½É±° œ¹Ý½É­™±½Üµ™½É´œ¤¹™½É…  ¡™½É´¤€ôøì(€½¹ÍÐ…ÉÑ¥±”€ô™½É´¹±½Í•ÍÐ …ÉÑ¥±”œ¤ì(€½¹ÍÐ™¥±•%¹ÁÕÐ€ô™½É´¹ÅÕ•ÉåM•±•Ñ½È ¥¹ÁÕÑmÑåÁ”ô‰™¥±”‰tœ¤ì(€½¹ÍÐÍÑ•À€ô9Õµ‰•È ¡™½É´¹…Ñ¥½¸¹µ…Ñ  ½p½ÍÑ•Áp¼¡q¬¤¼¤ñðmt¥lÅtñð€À¤ì(€½¹ÍÐ…•ÁÑÌ€ôì(€€€€ÄèíåÕ¹•áÁÉ•ÍÌè€œ¹ÍØœ°Í˜è€œ¹ÍØô°(€€€€ÈèíåÕ¹•áÁÉ•ÍÌè€œ¹á±Íàœ°Í˜è€œ¹á±Ì°¹á±Íàô°(€€€€ÌèíåÕ¹•áÁÉ•ÍÌè€œ¹é¥Àœ°Í˜è€œ¹Á‘˜ô(€ôì(€½¹ÍÐÍå¹A±…Ñ™½É´€ô€ ¤€ôøì(€€€½¹ÍÐÍ•±•Ñ•€ô™½É´¹ÅÕ•ÉåM•±•Ñ½È ¥¹ÁÕÑm¹…µ”ô‰Á±…Ñ™½É´‰té¡•­•œ¤ì(€€€½¹ÍÐÁ±…Ñ™½É´€ôÍ•±•Ñ•€üÍ•±•Ñ•¹Ù…±Õ”€è€åÕ¹•áÁÉ•ÍÌœì(€€€™½É´¹‘…Ñ…Í•Ð¹Á±…Ñ™½É´€ôÁ±…Ñ™½É´ì(€€€¥˜€¡…ÉÑ¥±”¤…ÉÑ¥±”¹‘…Ñ…Í•Ð¹Á±…Ñ™½É´€ôÁ±…Ñ™½É´ì(€€€¥˜€¡™¥±•%¹ÁÕÐ€˜˜…•ÁÑÍmÍÑ•Át¤ì(€€€€€™¥±•%¹ÁÕÐ¹…•ÁÐ€ô…•ÁÑÍmÍÑ•ÁumÁ±…Ñ™½Éµtì(€€€€€™¥±•%¹ÁÕÐ¹Ù…±Õ”€ô€œœì(€€€€€½¹ÍÐ…ÁÑ¥½¸€ô™¥±•%¹ÁÕÐ¹Á…É•¹Ñ±•µ•¹Ð¹ÅÕ•ÉåM•±•Ñ½È œéÍ½Á”€øÍÁ…¸é±…ÍÐµ¡¥±œ¤ì(€€€€€¥˜€¡…ÁÑ¥½¸¤…ÁÑ¥½¸¹Ñ•áÑ½¹Ñ•¹Ð€ô€Ÿ¦'š.§šZ’îØœì(€€€ô(€ôì(€™½É´¹ÅÕ•ÉåM•±•Ñ½É±° ¥¹ÁÕÑm¹…µ”ô‰Á±…Ñ™½É´‰tœ¤¹™½É…  ¡¥¹ÁÕÐ¤€ôøì(€€€¥¹ÁÕÐ¹…‘‘Ù•¹Ñ1¥ÍÑ•¹•È ¡…¹”œ°Íå¹A±…Ñ™½É´¤ì(€ô¤ì(€¥˜€¡™¥±•%¹ÁÕÐ¤ì(€€€™¥±•%¹ÁÕÐ¹…‘‘Ù•¹Ñ1¥ÍÑ•¹•È ¡…¹”œ°€ ¤€ôøì(€€€€€½¹ÍÐ…ÁÑ¥½¸€ô™¥±•%¹ÁÕÐ¹Á…É•¹Ñ±•µ•¹Ð¹ÅÕ•ÉåM•±•Ñ½È œéÍ½Á”€øÍÁ…¸é±…ÍÐµ¡¥±œ¤ì(€€€€€¥˜€¡…ÁÑ¥½¸¤…ÁÑ¥½¸¹Ñ•áÑ½¹Ñ•¹Ð€ô™¥±•%¹ÁÕÐ¹™¥±•Ì¹±•¹Ñ €ü™¥±•%¹ÁÕÐ¹™¥±•ÍlÁt¹¹…µ”€è€Ÿ¦'š.§šZ’îØœì(€€€ô¤ì(€ô(€Íå¹A±…Ñ™½É´ ¤ì)ô¤ì(œœœ(